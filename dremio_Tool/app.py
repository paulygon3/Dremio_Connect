"""
================================================================================
app.py - Main Application
================================================================================
Contains the main DremioExporter application class with the GUI.
================================================================================
"""

import codecs
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
from pathlib import Path
from datetime import datetime, date, time
import threading
import queue
import logging
import warnings
import weakref
import decimal
import math
from time import perf_counter
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# openpyxl's own compiled pattern for the bytes it refuses, rather than a second
# copy of the rule here. It leaves TAB, LF and CR alone, which is what Excel
# permits - see _sanitise_illegal_characters (F-05).
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from constants import (
    APP_TITLE, APP_SUBTITLE, APP_VERSION, COPYRIGHT,
    WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT,
    COLORS, DEFAULT_QUERY,
    EXCEL_MAX_CELL_CHARS, TRUNCATION_REPORT_LIMIT,
    EXCEL_MAX_DATA_ROWS, EXCEL_MAX_SHEET_ROWS, EXCEL_MAX_COLUMNS,
    ILLEGAL_CHAR_REPLACEMENT, SANITISED_REPORT_LIMIT,
    FORMULA_TRIGGER_CHARS, FORMULA_PREFIX,
    HISTORY_LABEL_LENGTH,
    QUERY_FILE_ENCODINGS, TEXT_FILE_ENCODING,
    DEFAULT_SHEET_NAME,
    UI_QUEUE_POLL_MS,
    DEFAULT_LOG_RETENTION_DAYS
)
from config import ConfigManager
from connection import DremioConnection, QueryCancelled
from utils import (
    load_logo_image, load_icon,
    validate_connection_params, validate_sheet_name,
    generate_timestamp_filename, validate_output_filename
)


class DremioExporter:
    """
    Main application class for Dremio to Excel Exporter.
    
    Single-page layout with:
        - Left panel: Connection settings, output options
        - Right panel: Query editor, execution log
        - Header: Logo, title, connection status
        - Status bar: Current state, version info
    """

    def __init__(self, root):
        """
        Initialize the application.
        
        Args:
            root: tkinter Tk() root window
        """
        self.root = root
        self.root.title(APP_TITLE)
        self.root.configure(bg=COLORS['white'])
        
        # Initialize managers
        self.config = ConfigManager()
        self._setup_file_logging()
        self.connection = DremioConnection()
        
        # Load window size from config
        width = self.config.get('ui', 'window_width', 1100)
        height = self.config.get('ui', 'window_height', 800)
        self.root.geometry(f"{width}x{height}")
        self.root.minsize(WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT)
        
        # Set window icon
        self._set_window_icon()
        
        # Load logo image
        self.logo_image = load_logo_image()
        
        # State variables
        self.is_running = False
        # Set by _stop_execution, read by the worker and passed down to
        # connection.execute_query, which checks it between record batches.
        # An Event rather than a bool because it crosses threads (F-13).
        self.cancel_requested = threading.Event()
        self.df = None
        # Populated by _export_to_excel when cells had to be truncated; read by
        # _execute_thread to tell the user. Reset at the start of every export.
        self.last_truncation = None
        # Any other warning raised during the write. Previously these went to
        # stderr, which a windowed build does not have.
        self.last_export_warnings = []
        # Cells whose control characters had to be removed for openpyxl to
        # accept them (F-05), as (sheet_row, column, removed_count).
        self.last_sanitised = []
        # Cells that began with a formula character and were quoted so Excel
        # keeps them as text (CWE-1236), as (sheet_row, column).
        self.last_neutralised = []
        # Set by _on_close before the root is destroyed. Read by _ui() so a
        # worker that is still unwinding stops trying to reach a UI that no
        # longer exists (F-12).
        self.shutting_down = False
        # Worker threads hand UI work to the Tk thread through this queue
        # rather than calling root.after themselves. See _ui.
        self._ui_queue = queue.Queue()
        
        # Build UI
        self._setup_styles()
        self._create_header()
        self._create_main_content()
        self._create_status_bar()
        
        # Load saved settings
        self._load_saved_settings()

        if self.log_file_path:
            self._log(f"Session log: {self.log_file_path}")

        # Bind events
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Start the pump that delivers worker updates to the UI (F-12).
        self.root.after(UI_QUEUE_POLL_MS, self._drain_ui_queue)
    
    def _set_window_icon(self):
        """Set the window icon from auto-detected .ico file."""
        load_icon(self.root)
    
    def _setup_styles(self):
        """Configure ttk widget styles for professional appearance."""
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # LabelFrame - clean card style with subtle border
        self.style.configure('Card.TLabelframe', 
                           background=COLORS['white'],
                           borderwidth=1,
                           relief='solid',
                           bordercolor=COLORS['border'])
        self.style.configure('Card.TLabelframe.Label', 
                           font=('Segoe UI', 10, 'bold'),
                           foreground=COLORS['header_bg'],
                           background=COLORS['white'])
        
        # Entry fields - subtle border
        self.style.configure('TEntry', 
                           padding=5,
                           relief='solid',
                           borderwidth=1)
        self.style.map('TEntry',
                      bordercolor=[('focus', COLORS['header_bg']), 
                                   ('!focus', COLORS['border'])])
        
        # Combobox
        self.style.configure('TCombobox', 
                           padding=5,
                           relief='solid',
                           borderwidth=1)
        
        # Checkbutton
        self.style.configure('TCheckbutton',
                           background=COLORS['white'],
                           font=('Segoe UI', 9))
        
        # Buttons - flat modern style
        self.style.configure('TButton',
                           font=('Segoe UI', 9),
                           padding=(10, 5),
                           relief='flat',
                           borderwidth=0)
        self.style.map('TButton',
                      background=[('active', COLORS['border']),
                                  ('!active', COLORS['frame_bg'])])
    
    # =========================================================================
    # HEADER
    # =========================================================================
    
    def _create_header(self):
        """Create header with logo, title, and connection status."""
        header = tk.Frame(self.root, bg=COLORS['header_bg'], height=70)
        header.pack(fill='x')
        header.pack_propagate(False)
        
        # Logo area - always reserve space for consistent layout
        if self.logo_image:
            # Show actual image
            logo_label = tk.Label(header, image=self.logo_image, 
                                 bg=COLORS['header_bg'])
            logo_label.pack(side='left', padx=15, pady=12)
        else:
            # Blank placeholder - maintains layout spacing
            logo_placeholder = tk.Frame(header, bg=COLORS['header_bg'], 
                                        width=45, height=45)
            logo_placeholder.pack(side='left', padx=15, pady=12)
            logo_placeholder.pack_propagate(False)
        
        # Title
        title_frame = tk.Frame(header, bg=COLORS['header_bg'])
        title_frame.pack(side='left', pady=12)
        
        tk.Label(title_frame, text=APP_TITLE,
                font=('Segoe UI', 16, 'bold'),
                fg=COLORS['header_fg'],
                bg=COLORS['header_bg']).pack(anchor='w')
        
        tk.Label(title_frame, text=APP_SUBTITLE,
                font=('Segoe UI', 9),
                fg=COLORS['header_subtitle'],
                bg=COLORS['header_bg']).pack(anchor='w')
        
        # Connection status (right side)
        status_frame = tk.Frame(header, bg=COLORS['header_bg'])
        status_frame.pack(side='right', padx=20, pady=12)
        
        self.conn_indicator = tk.Label(status_frame, text="●",
                                       font=('Segoe UI', 16),
                                       fg=COLORS['disconnected'],
                                       bg=COLORS['header_bg'])
        self.conn_indicator.pack(side='left')
        
        self.conn_label = tk.Label(status_frame, text="Not Connected",
                                   font=('Segoe UI', 10),
                                   fg=COLORS['header_fg'],
                                   bg=COLORS['header_bg'])
        self.conn_label.pack(side='left', padx=(5, 0))
    
    # =========================================================================
    # MAIN CONTENT
    # =========================================================================
    
    def _create_main_content(self):
        """Create main two-panel layout."""
        main = tk.Frame(self.root, bg=COLORS['frame_bg'])
        main.pack(fill='both', expand=True, padx=15, pady=10)
        
        # Left panel - Connection & Output
        left_panel = tk.Frame(main, bg=COLORS['frame_bg'], width=320)
        left_panel.pack(side='left', fill='y', padx=(0, 10))
        left_panel.pack_propagate(False)
        
        self._create_connection_panel(left_panel)
        self._create_output_panel(left_panel)
        
        # Right panel - Query & Log
        right_panel = tk.Frame(main, bg=COLORS['frame_bg'])
        right_panel.pack(side='left', fill='both', expand=True)
        
        self._create_query_panel(right_panel)
        self._create_log_panel(right_panel)
    
    def _create_connection_panel(self, parent):
        """Create connection settings panel."""
        frame = ttk.LabelFrame(parent, text="Connection", style='Card.TLabelframe')
        frame.pack(fill='x', pady=(0, 10))
        
        inner = tk.Frame(frame, bg=COLORS['white'])
        inner.pack(fill='x', padx=10, pady=10)
        
        # Fields
        fields = [
            ("Hostname:", "hostname", False),
            ("Port:", "port", False),
            ("Username:", "username", False),
            ("PAT Token:", "token", True),
        ]
        
        self.conn_fields = {}
        for i, (label, name, is_password) in enumerate(fields):
            tk.Label(inner, text=label, font=('Segoe UI', 9),
                    bg=COLORS['white']).grid(row=i, column=0, sticky='w', pady=3)
            
            entry = ttk.Entry(inner, width=28, show="•" if is_password else "")
            entry.grid(row=i, column=1, sticky='ew', pady=3)
            self.conn_fields[name] = entry
        
        # Username change handler (to load saved token)
        self.conn_fields['username'].bind('<FocusOut>', self._on_username_change)
        
        inner.columnconfigure(1, weight=1)
        
        # Options
        row = len(fields)
        
        self.remember_token = tk.BooleanVar(value=True)
        ttk.Checkbutton(inner, text="Remember token",
                       variable=self.remember_token).grid(
            row=row, column=0, columnspan=2, sticky='w', pady=(5, 0))
        
        self.use_tls = tk.BooleanVar(value=True)
        ttk.Checkbutton(inner, text="Use TLS/SSL",
                       variable=self.use_tls).grid(
            row=row+1, column=0, columnspan=2, sticky='w')
        
        # Connect button
        btn_frame = tk.Frame(inner, bg=COLORS['white'])
        btn_frame.grid(row=row+2, column=0, columnspan=2, sticky='ew', pady=(10, 0))
        
        self.connect_btn = tk.Button(
            btn_frame, text="Connect",
            font=('Segoe UI', 9, 'bold'),
            bg=COLORS['button_primary'], fg='white',
            activebackground='#004C8C', activeforeground='white',
            relief='flat', padx=15, pady=6, cursor='hand2',
            command=self._toggle_connection
        )
        self.connect_btn.pack(fill='x')
    
    def _create_output_panel(self, parent):
        """Create output settings panel."""
        frame = ttk.LabelFrame(parent, text="Output Settings", style='Card.TLabelframe')
        frame.pack(fill='x', pady=(0, 10))
        
        inner = tk.Frame(frame, bg=COLORS['white'])
        inner.pack(fill='x', padx=10, pady=10)
        
        # Output directory
        tk.Label(inner, text="Output Folder:", font=('Segoe UI', 9),
                bg=COLORS['white']).pack(anchor='w')
        
        dir_frame = tk.Frame(inner, bg=COLORS['white'])
        dir_frame.pack(fill='x', pady=(2, 8))
        
        self.output_dir = ttk.Entry(dir_frame, width=22)
        self.output_dir.pack(side='left', fill='x', expand=True)
        
        ttk.Button(dir_frame, text="...", width=3,
                  command=self._browse_output).pack(side='left', padx=(5, 0))
        
        # Filename
        tk.Label(inner, text="Filename:", font=('Segoe UI', 9),
                bg=COLORS['white']).pack(anchor='w')
        self.filename = ttk.Entry(inner)
        self.filename.pack(fill='x', pady=(2, 8))
        
        # Options
        self.open_after = tk.BooleanVar(value=True)
        ttk.Checkbutton(inner, text="Open after export",
                       variable=self.open_after).pack(anchor='w')
        
        self.autofit = tk.BooleanVar(value=True)
        ttk.Checkbutton(inner, text="Auto-fit columns",
                       variable=self.autofit).pack(anchor='w')
        
        self.freeze_header = tk.BooleanVar(value=True)
        ttk.Checkbutton(inner, text="Freeze header row",
                       variable=self.freeze_header).pack(anchor='w')

        # Log retention (saved). 0 = keep every session log forever.
        retention_row = tk.Frame(inner, bg=COLORS['white'])
        retention_row.pack(anchor='w', fill='x', pady=(6, 0))
        tk.Label(retention_row, text="Keep logs (days, 0 = forever):",
                 font=('Segoe UI', 9), bg=COLORS['white']).pack(side='left')
        self.log_retention_days = tk.IntVar(value=DEFAULT_LOG_RETENTION_DAYS)
        ttk.Spinbox(retention_row, from_=0, to=3650, width=6,
                    textvariable=self.log_retention_days).pack(side='left', padx=(6, 0))
    
    def _create_query_panel(self, parent):
        """Create query editor panel."""
        frame = ttk.LabelFrame(parent, text="SQL Query", style='Card.TLabelframe')
        frame.pack(fill='both', expand=True, pady=(0, 10))
        
        inner = tk.Frame(frame, bg=COLORS['white'])
        inner.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Toolbar
        toolbar = tk.Frame(inner, bg=COLORS['white'])
        toolbar.pack(fill='x', pady=(0, 5))
        
        tk.Label(toolbar, text="Recent:", font=('Segoe UI', 9),
                bg=COLORS['white']).pack(side='left')
        
        # Width tied to the label bound, not a separate number. At width=40 a
        # 60-character label was clipped at exactly the point that distinguishes
        # two entries, which put F-27 back on screen after the labels themselves
        # had been fixed. Tying the two means they cannot drift apart again.
        self.history_combo = ttk.Combobox(toolbar, width=HISTORY_LABEL_LENGTH,
                                          state='readonly')
        self.history_combo.pack(side='left', padx=(5, 10))
        self.history_combo.bind('<<ComboboxSelected>>', self._load_from_history)
        self._update_history_dropdown()
        
        # "Library" rather than "Load": these now open the managed query
        # collection in saved_queries/ (F-31), which config.py has always
        # implemented and nothing ever called. Opening an arbitrary .sql file is
        # still available, from inside that dialog.
        ttk.Button(toolbar, text="Library", command=self._open_query_library,
                  width=8).pack(side='left', padx=2)
        ttk.Button(toolbar, text="Save", command=self._save_query_file,
                  width=6).pack(side='left', padx=2)
        ttk.Button(toolbar, text="Clear", command=self._clear_query,
                  width=6).pack(side='left', padx=2)
        
        # Query text editor with clean border
        query_frame = tk.Frame(inner, bg=COLORS['border'], padx=1, pady=1)
        query_frame.pack(fill='both', expand=True)
        
        self.query_text = scrolledtext.ScrolledText(
            query_frame, font=('Consolas', 10),
            wrap='none', height=12,
            relief='flat', borderwidth=0,
            highlightthickness=0
        )
        self.query_text.pack(fill='both', expand=True)
        
        # Execute buttons
        exec_frame = tk.Frame(inner, bg=COLORS['white'])
        exec_frame.pack(fill='x', pady=(10, 0))
        
        self.execute_btn = tk.Button(
            exec_frame, text="▶  Execute Query and Export to Excel",
            font=('Segoe UI', 11, 'bold'),
            bg=COLORS['button_success'], fg='white',
            activebackground='#1E7E34', activeforeground='white',
            relief='flat', padx=20, pady=10, cursor='hand2',
            command=self._execute_and_export,
            state='disabled'
        )
        self.execute_btn.pack(side='left')
        
        self.stop_btn = tk.Button(
            exec_frame, text="■ Stop",
            font=('Segoe UI', 10),
            bg=COLORS['button_secondary'], fg='white',
            relief='flat', padx=15, pady=10,
            state='disabled', command=self._stop_execution
        )
        self.stop_btn.pack(side='left', padx=(10, 0))
        
        self.progress_label = tk.Label(
            exec_frame, text="", font=('Segoe UI', 9),
            fg=COLORS['text_muted'], bg=COLORS['white']
        )
        self.progress_label.pack(side='right')
    
    def _create_log_panel(self, parent):
        """Create execution log panel."""
        frame = ttk.LabelFrame(parent, text="Execution Log", style='Card.TLabelframe')
        frame.pack(fill='both', expand=True)
        
        inner = tk.Frame(frame, bg=COLORS['white'])
        inner.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Log text with clean border
        log_frame = tk.Frame(inner, bg=COLORS['border'], padx=1, pady=1)
        log_frame.pack(fill='both', expand=True)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, font=('Consolas', 9),
            bg=COLORS['log_bg'], fg=COLORS['log_fg'],
            height=8, relief='flat', borderwidth=0,
            highlightthickness=0
        )
        self.log_text.pack(fill='both', expand=True)
        
        # Buttons
        btn_frame = tk.Frame(inner, bg=COLORS['white'])
        btn_frame.pack(fill='x', pady=(5, 0))
        
        ttk.Button(btn_frame, text="Clear Log",
                  command=self._clear_log, width=10).pack(side='left')
        ttk.Button(btn_frame, text="Save Log",
                  command=self._save_log, width=10).pack(side='left', padx=(5, 0))
        ttk.Button(btn_frame, text="Open Output Folder",
                  command=self._open_output_folder).pack(side='right')
        
        # Initial log
        self._log(f"Application started - v{APP_VERSION}")

        # Anything ConfigManager found wrong on load. It runs before any widget
        # exists, so it collects its complaints and they are drained here, at
        # the first moment there is somewhere to show them.
        for warning in self.config.load_warnings:
            self._log(f"WARNING: {warning}")

        # From here on it can tell us directly. print() was its only error
        # channel, and a windowed build has no console for those to land in
        # (F-20). _ui, not _log: saves also happen on worker threads.
        #
        # Through a weakref, because a closure over self would complete the
        # cycle app -> config -> callback -> app. A cycle is collected by the
        # cyclic GC rather than by refcount, and that can run on ANY thread -
        # which finalises this app's Tk variables off the Tk thread and aborts
        # the process with "Tcl_AsyncDelete: async handler deleted by the wrong
        # thread". Observed, not hypothesised: it killed two repro scripts.
        weak_self = weakref.ref(self)
        def report_config_warning(message):
            app = weak_self()
            if app is not None:
                app._ui(lambda: app._log(f"WARNING: {message}"))
        self.config.on_warning = report_config_warning

        self._log("Enter credentials and click Connect")
    
    # =========================================================================
    # STATUS BAR
    # =========================================================================
    
    def _create_status_bar(self):
        """Create bottom status bar."""
        status = tk.Frame(self.root, bg='#E9ECEF', height=25)
        status.pack(fill='x', side='bottom')
        status.pack_propagate(False)
        
        self.status_text = tk.Label(
            status, text="Ready", font=('Segoe UI', 9),
            fg=COLORS['text_muted'], bg='#E9ECEF'
        )
        self.status_text.pack(side='left', padx=10)
        
        tk.Label(status, text=COPYRIGHT,
                font=('Segoe UI', 9),
                fg=COLORS['text_muted'],
                bg='#E9ECEF').pack(side='right', padx=10)
    
    # =========================================================================
    # SETTINGS MANAGEMENT
    # =========================================================================
    
    def _load_saved_settings(self):
        """Load saved settings into form fields."""
        # Connection
        self.conn_fields['hostname'].insert(0, self.config.get('connection', 'hostname', ''))
        self.conn_fields['port'].insert(0, self.config.get('connection', 'port', '32010'))
        self.conn_fields['username'].insert(0, self.config.get('connection', 'username', ''))
        self.use_tls.set(self.config.get('connection', 'use_tls', True))
        
        # Load saved token
        username = self.config.get('connection', 'username', '')
        if username:
            saved_token = self.config.get_token(username)
            if saved_token:
                self.conn_fields['token'].insert(0, saved_token)
        
        # Output
        default_dir = str(Path.home() / 'Documents' / 'Dremio_Exports')
        self.output_dir.insert(0, self.config.get('output', 'directory', default_dir))
        self.filename.insert(0, self.config.get('output', 'filename_pattern', 
                                                'dremio_export_{timestamp}.xlsx'))
        self.open_after.set(self.config.get('output', 'open_after_export', True))
        self.autofit.set(self.config.get('output', 'autofit_columns', True))
        self.freeze_header.set(self.config.get('output', 'freeze_header', True))
        self.log_retention_days.set(
            self.config.get('logging', 'log_retention_days',
                            DEFAULT_LOG_RETENTION_DAYS))
        
        # Query
        last_query = self.config.get('ui', 'last_query', '')
        self.query_text.insert('1.0', last_query if last_query else DEFAULT_QUERY)
    
    def _save_current_settings(self):
        """Save current settings to config."""
        # Connection
        self.config.set('connection', 'hostname', self.conn_fields['hostname'].get())
        self.config.set('connection', 'port', self.conn_fields['port'].get())
        self.config.set('connection', 'username', self.conn_fields['username'].get())
        self.config.set('connection', 'use_tls', self.use_tls.get())
        
        # Output
        self.config.set('output', 'directory', self.output_dir.get())
        self.config.set('output', 'filename_pattern', self.filename.get())
        self.config.set('output', 'open_after_export', self.open_after.get())
        self.config.set('output', 'autofit_columns', self.autofit.get())
        self.config.set('output', 'freeze_header', self.freeze_header.get())

        # Logging
        try:
            retention = int(self.log_retention_days.get())
        except (tk.TclError, ValueError):
            retention = DEFAULT_LOG_RETENTION_DAYS
        self.config.set('logging', 'log_retention_days', max(0, retention))
        
        # UI
        self.config.set('ui', 'window_width', self.root.winfo_width())
        self.config.set('ui', 'window_height', self.root.winfo_height())
        self.config.set('ui', 'last_query', self.query_text.get('1.0', 'end-1c'))
        
        # Token
        username = self.conn_fields['username'].get()
        if self.remember_token.get():
            token = self.conn_fields['token'].get()
            if username and token:
                self.config.save_token(username, token)
        elif username:
            # Unchecking the box is the only control the UI offers for
            # forgetting a stored credential, so it has to actually forget it.
            # Without this branch the PAT stayed on disk and
            # _on_username_change read it straight back out on the next launch.
            self.config.delete_token(username)
        
        self.config.save_config()
    
    def _on_username_change(self, event=None):
        """Load saved token when username changes."""
        username = self.conn_fields['username'].get()
        if username and not self.conn_fields['token'].get():
            saved_token = self.config.get_token(username)
            if saved_token:
                self.conn_fields['token'].delete(0, 'end')
                self.conn_fields['token'].insert(0, saved_token)
    
    def _on_close(self):
        """
        Handle application close.

        Closing the window while a query is running used to strand the worker:
        it kept reading, then blocked forever on its first Tk call into a
        destroyed interpreter (F-12). Two things prevent that, in this order.

        Cancel first, so the worker is not sitting inside a multi-second read
        when the interpreter disappears - it unwinds through the same
        cancellation path as Stop. Then set the flag before destroy(), so every
        remaining update the worker tries to marshal is dropped rather than
        queued against an interpreter that is going away.
        """
        if self.is_running:
            self.cancel_requested.set()
            self.connection.cancel_query()

        self._save_current_settings()
        logger = getattr(self, 'logger', None)
        if logger is not None:
            logger.info("=== Session ended ===")
            for h in list(logger.handlers):
                h.close()
                logger.removeHandler(h)
        self.shutting_down = True
        self.root.destroy()
    
    # =========================================================================
    # UI HELPERS
    # =========================================================================
    
    def _ui(self, fn):
        """
        Hand `fn` to the Tk thread. Safe to call from any thread.

        Every cross-thread update used to go through root.after(0, ...)
        directly from the worker. That is the documented primitive, and it is
        still fine while the window is open - but it has no answer for the
        window closing mid-query. root.after from a foreign thread enters Tcl,
        which queues the call to the interpreter's own thread and BLOCKS until
        that thread runs it. Close the window at that moment and nothing ever
        runs it: the worker is stranded inside Tcl forever, the export is
        abandoned, and no error appears on any channel (F-12, measured - the
        worker was blocked in Tkinter's _register).

        A flag cannot fix that, because the call is already inside Tcl before
        any flag can be read. So the worker does not enter Tcl at all: it puts
        a callable on a plain queue, and _drain_ui_queue - which runs on the Tk
        thread, where touching Tcl is legal - executes it.

        Returns:
            bool: True if the callback was accepted, False if the UI is gone.
        """
        if self.shutting_down:
            return False
        self._ui_queue.put(fn)
        return True

    def _drain_ui_queue(self):
        """
        Run whatever the workers have queued, then reschedule.

        Tk thread only. Each callback is isolated: one that raises must not
        stop the pump, or every later update - including the ones reporting
        that failure - would be lost silently.
        """
        while True:
            try:
                callback = self._ui_queue.get_nowait()
            except queue.Empty:
                break
            if self.shutting_down:
                break
            try:
                callback()
            except Exception:
                # Same destination an after() callback's exception would have
                # reached, so this does not change where errors surface.
                try:
                    self.root.report_callback_exception(*sys.exc_info())
                except Exception:
                    pass

        if not self.shutting_down:
            self.root.after(UI_QUEUE_POLL_MS, self._drain_ui_queue)

    def _setup_file_logging(self):
        """
        Open a per-session .txt log under the app-data folder.

        Every line shown in the log panel is mirrored here (see _log), together
        with the lower-level connection and RPC status, so there is a durable
        record after the window closes - a windowed build has no console to
        scroll back through. Best-effort: the panel still works if this fails.
        """
        self.log_file_path = None
        self.logger = None
        try:
            logs_dir = self.config.app_dir / 'logs'
            logs_dir.mkdir(exist_ok=True)
            retention = self.config.get(
                'logging', 'log_retention_days', DEFAULT_LOG_RETENTION_DAYS)
            self._prune_old_logs(logs_dir, retention)
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.log_file_path = logs_dir / f'dremio_log_{stamp}.txt'
            logger = logging.getLogger(f'dremio_exporter.{id(self)}')
            logger.setLevel(logging.INFO)
            logger.propagate = False
            handler = logging.FileHandler(self.log_file_path, encoding='utf-8')
            handler.setFormatter(logging.Formatter(
                '%(asctime)s %(message)s', '%Y-%m-%d %H:%M:%S'))
            logger.addHandler(handler)
            logger.info(f"=== Session started (v{APP_VERSION}) ===")
            self.logger = logger
        except Exception:
            self.logger = None

    def _prune_old_logs(self, logs_dir, retention_days):
        """
        Delete session logs older than retention_days; 0 keeps them forever.

        Best-effort and quiet - a log that will not delete must not stop the
        app from starting.
        """
        try:
            days = int(retention_days)
        except (TypeError, ValueError):
            days = DEFAULT_LOG_RETENTION_DAYS
        if days <= 0:
            return
        cutoff = datetime.now().timestamp() - days * 86400
        for path in logs_dir.glob('dremio_log_*.txt'):
            try:
                if path.stat().st_mtime < cutoff:
                    path.unlink()
            except OSError:
                pass

    def _log(self, message):
        """
        Add message to log.

        update_idletasks, not update: update() dispatches the ENTIRE pending
        queue - other after callbacks and user input alike - so every log line
        re-entered the event loop while its caller was still on the stack.
        Queued callbacks ran out of order, a click on Disconnect could re-enter
        the connection handler mid-log, and closing the window mid-log left the
        caller resuming against destroyed widgets (F-15). update_idletasks
        flushes the redraw, which is all this needs, and dispatches nothing.
        """
        # Mirror every panel line to the durable session .txt log, before the
        # shutting-down guard so shutdown lines are captured too.
        logger = getattr(self, 'logger', None)
        if logger is not None:
            logger.info(message)
        if self.shutting_down:
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert('end', f"{timestamp} - {message}\n")
        self.log_text.see('end')
        self.root.update_idletasks()
    
    def _clear_log(self):
        """Clear the log."""
        self.log_text.delete('1.0', 'end')
    
    def _write_text_file(self, filepath, text, what):
        """
        Write a text file from a UI callback, reporting any failure.

        Unguarded, these ran in Tk button callbacks with no handler, so an
        OSError - a read-only target, a full disk, a path on a disconnected
        share - propagated into Tk, which printed a traceback to stderr and
        carried on. The button appeared to do nothing, and a windowed
        PyInstaller build has no stderr, so it was completely silent (F-18).

        The encoding is explicit for the same reason it is in config.py: the
        platform default is not the same thing on every machine, and this app
        wrote its own files one way and read them another.

        Returns:
            bool: True if the file was written.
        """
        try:
            with open(filepath, 'w', encoding=TEXT_FILE_ENCODING) as f:
                f.write(text)
        except OSError as e:
            self._log(f"ERROR: could not save the {what}: {e}")
            messagebox.showerror(
                "Could Not Save",
                f"The {what} could not be saved to:\n{filepath}\n\n{e}\n\n"
                f"Check that the folder exists and is writable, then try again."
            )
            return False

        self._log(f"{what.capitalize()} saved: {Path(filepath).name}")
        return True

    def _save_log(self):
        """Save log to file."""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".txt",
            initialfile=f"dremio_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        if filepath:
            self._write_text_file(filepath, self.log_text.get('1.0', 'end'),
                                  "log")
    
    def _clear_query(self):
        """Clear query editor."""
        self.query_text.delete('1.0', 'end')
    
    def _read_query_file(self, filepath):
        """
        Read a .sql file the user picked, whatever it happens to be encoded as.

        The old call passed no encoding, so it got the platform default and
        raised UnicodeDecodeError on anything else - straight into the Tk
        callback, with no handler and nothing shown (F-18). That is not an
        exotic failure: .sql files come out of SSMS as UTF-16 and out of older
        tools as cp1252, and neither decodes as UTF-8.

        The file is read once as bytes and decoded in memory rather than
        reopened per attempt, so a file that changes underneath cannot produce a
        mixture of two reads.

        A BOM is decisive, and that matters more than it looks. Trying UTF-16 as
        one option among several does not work: with no BOM it assumes
        little-endian and decodes ANY even-length byte sequence into
        plausible-looking nonsense rather than raising, so a cp1252 file came
        back as garbage instead of falling through to the encoding that could
        read it. An encoding that never fails cannot be part of a
        try-in-order list - it ends the list.

        Returns:
            tuple: (text, encoding_used, error). On success error is None; on
                failure text and encoding_used are None.
        """
        try:
            raw = Path(filepath).read_bytes()
        except OSError as e:
            return None, None, f"{e}"

        if raw.startswith(codecs.BOM_UTF8):
            candidates = ('utf-8-sig',)
        elif raw.startswith((codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)):
            candidates = ('utf-16',)
        else:
            candidates = QUERY_FILE_ENCODINGS

        for encoding in candidates:
            try:
                return raw.decode(encoding), encoding, None
            except (UnicodeDecodeError, LookupError):
                continue

        return None, None, (
            f"the file is not text in any encoding this app reads "
            f"({', '.join(candidates)}). If it is a SQL script, re-save it as "
            f"UTF-8 and try again."
        )

    def _put_query_in_editor(self, text, source):
        """Replace the editor contents and say where they came from."""
        self.query_text.delete('1.0', 'end')
        self.query_text.insert('1.0', text)
        self._log(f"Loaded: {source}")

    def _library_load(self, filepath):
        """
        Load one saved query into the editor. Returns True on success.

        Separated from the dialog so it can be driven - and tested - without
        one. config.load_query_file is the implementation that has been sitting
        unused since the beginning (F-31); this is the caller it never had.
        """
        try:
            text = self.config.load_query_file(filepath)
        except OSError as e:
            self._log(f"ERROR: could not open {Path(filepath).name}: {e}")
            messagebox.showerror(
                "Could Not Open",
                f"This saved query could not be read:\n{filepath}\n\n{e}"
            )
            return False
        except UnicodeDecodeError as e:
            # The library writes UTF-8, so this means the file was replaced or
            # hand-edited with something else. Fall back to the same detection
            # the Browse path uses rather than refusing outright (F-18).
            text, encoding, error = self._read_query_file(filepath)
            if error:
                self._log(f"ERROR: could not read {Path(filepath).name}: {error}")
                messagebox.showerror(
                    "Could Not Open",
                    f"This saved query could not be read:\n{filepath}\n\n{error}"
                )
                return False
            self._log(f"Note: {Path(filepath).name} was not UTF-8; "
                      f"decoded as {encoding}")

        self._put_query_in_editor(text, Path(filepath).name)
        return True

    def _library_delete(self, filepath):
        """Delete one saved query, after asking. Returns True if it went."""
        name = Path(filepath).name
        if not messagebox.askyesno(
                "Delete Saved Query",
                f"Delete '{name}' permanently?\n\nThis cannot be undone."):
            return False

        try:
            self.config.delete_query_file(filepath)
        except OSError as e:
            self._log(f"ERROR: could not delete {name}: {e}")
            messagebox.showerror("Could Not Delete",
                                 f"'{name}' could not be deleted:\n\n{e}")
            return False

        self._log(f"Deleted saved query: {name}")
        return True

    def _open_query_library(self):
        """
        Show the saved-query library, with an escape hatch to the filesystem.

        config.py has always implemented a saved-queries subsystem -
        get_saved_queries, save_query_file, load_query_file, delete_query_file -
        over a saved_queries/ directory it creates on every start. The UI never
        called any of it, using its own filedialog instead, so the app carried
        two implementations of one feature and shipped the one with no library
        (F-31).

        Browse is kept deliberately. The library is now the primary path, but
        opening a .sql file from anywhere was a documented capability and
        removing it to tidy up would be a regression dressed as a fix.
        """
        queries = self.config.get_saved_queries()

        dialog = tk.Toplevel(self.root)
        dialog.title("Query Library")
        dialog.transient(self.root)
        dialog.geometry("460x320")
        dialog.configure(bg=COLORS['white'])
        self.query_library_dialog = dialog

        tk.Label(dialog, text="Saved queries", bg=COLORS['white'],
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w', padx=12,
                                                     pady=(12, 4))

        listbox = tk.Listbox(dialog, activestyle='none')
        listbox.pack(fill='both', expand=True, padx=12)
        for path in queries:
            listbox.insert('end', path.stem)
        if queries:
            listbox.selection_set(0)
        else:
            listbox.insert('end', "(no saved queries yet - use Save)")
            listbox.config(state='disabled')

        buttons = tk.Frame(dialog, bg=COLORS['white'])
        buttons.pack(fill='x', padx=12, pady=12)

        def selected_path():
            if not queries:
                return None
            selection = listbox.curselection()
            return queries[selection[0]] if selection else None

        def do_load():
            path = selected_path()
            if path and self._library_load(path):
                dialog.destroy()

        def do_delete():
            path = selected_path()
            if path and self._library_delete(path):
                dialog.destroy()
                self._open_query_library()

        def do_browse():
            dialog.destroy()
            self._browse_query_file()

        ttk.Button(buttons, text="Open", command=do_load).pack(side='left')
        ttk.Button(buttons, text="Delete", command=do_delete).pack(side='left',
                                                                   padx=6)
        ttk.Button(buttons, text="Browse...", command=do_browse).pack(side='left')
        ttk.Button(buttons, text="Cancel",
                   command=dialog.destroy).pack(side='right')

        listbox.bind('<Double-Button-1>', lambda e: do_load())
        return dialog

    def _browse_query_file(self):
        """Open a .sql file from anywhere on disk."""
        filepath = filedialog.askopenfilename(
            filetypes=[("SQL Files", "*.sql"), ("All Files", "*.*")]
        )
        if not filepath:
            return

        text, encoding, error = self._read_query_file(filepath)
        if error:
            self._log(f"ERROR: could not load {Path(filepath).name}: {error}")
            messagebox.showerror(
                "Could Not Load",
                f"This file could not be read:\n{filepath}\n\n{error}"
            )
            return

        self.query_text.delete('1.0', 'end')
        self.query_text.insert('1.0', text)

        # Say so when it was not UTF-8. The text has been converted on the way
        # in, and the user is the only one who can tell whether the result is
        # still the query they meant.
        if encoding != QUERY_FILE_ENCODINGS[0]:
            self._log(f"Loaded: {Path(filepath).name} (decoded as {encoding})")
        else:
            self._log(f"Loaded: {Path(filepath).name}")
    
    def _save_query_file(self):
        """
        Save the current query into the library, by name.

        config.save_query_file has always existed and was never called (F-31).
        Asking for a name rather than a path is what makes the collection a
        library instead of a scatter of files: everything lands in
        saved_queries/, which is the directory the app already creates on every
        start and never used.
        """
        query = self.query_text.get('1.0', 'end-1c').strip()
        if not query:
            messagebox.showwarning("Nothing to Save",
                                   "The query editor is empty.")
            return

        name = simpledialog.askstring(
            "Save Query",
            "Name for this query:",
            parent=self.root,
        )
        if name is None:
            return

        # Ask before replacing, for the reason F-25 gives about exports: the
        # library is a list of names, so a collision is easy to make by accident
        # and silent overwriting loses work with nothing said.
        safe_name = self.config.clean_query_name(name)
        if safe_name:
            existing = self.config.queries_dir / f"{safe_name}.sql"
            if existing.exists() and not messagebox.askyesno(
                    "Replace Saved Query",
                    f"'{safe_name}' already exists.\n\nReplace it?"):
                return

        try:
            filepath = self.config.save_query_file(name, query)
        except ValueError as e:
            self._log(f"ERROR: {e}")
            messagebox.showerror("Invalid Name", str(e))
            return
        except OSError as e:
            self._log(f"ERROR: could not save the query: {e}")
            messagebox.showerror(
                "Could Not Save",
                f"'{name}' could not be saved:\n\n{e}"
            )
            return

        self._log(f"Saved to the query library: {filepath.name}")
    
    def _browse_output(self):
        """Browse for output directory."""
        folder = filedialog.askdirectory()
        if folder:
            self.output_dir.delete(0, 'end')
            self.output_dir.insert(0, folder)
    
    def _open_output_folder(self):
        """
        Open output folder in file explorer.

        Every failure here is reported. Unguarded, this ran in a Tk button
        callback with no handler: Tk caught the exception, printed a traceback
        to stderr and carried on, so the button simply appeared to do nothing -
        and in a windowed PyInstaller build there is no stderr, making it
        completely silent.
        """
        folder = Path(self.output_dir.get())

        if not folder.exists():
            messagebox.showwarning("Not Found", f"Folder does not exist:\n{folder}")
            return

        if not hasattr(os, 'startfile'):
            self._log(f"Opening a folder is supported on Windows only: {folder}")
            messagebox.showinfo(
                "Not Supported",
                f"Opening a folder from the app is supported on Windows only.\n\n"
                f"The folder is:\n{folder}"
            )
            return

        try:
            os.startfile(folder)
        except Exception as e:
            self._log(f"ERROR: could not open the output folder: {e}")
            messagebox.showerror(
                "Could Not Open Folder", f"{folder}\n\n{e}"
            )
    
    def _update_history_dropdown(self):
        """Update query history dropdown."""
        labels = self.config.get_history_labels()
        self.history_combo['values'] = labels if labels else ['(No recent queries)']
    
    def _load_from_history(self, event=None):
        """Load query from history."""
        idx = self.history_combo.current()
        query = self.config.get_query_from_history(idx)
        if query:
            self.query_text.delete('1.0', 'end')
            self.query_text.insert('1.0', query)
            self._log("Loaded query from history")
    
    def _update_connection_status(self, connected, server=""):
        """Update connection indicator in header."""
        if connected:
            self.conn_indicator.config(fg=COLORS['connected'])
            self.conn_label.config(text=f"Connected: {server}")
            self.connect_btn.config(text="Disconnect", bg=COLORS['button_danger'])
            self.execute_btn.config(state='normal')
        else:
            self.conn_indicator.config(fg=COLORS['disconnected'])
            self.conn_label.config(text="Not Connected")
            self.connect_btn.config(text="Connect", bg=COLORS['button_primary'])
            self.execute_btn.config(state='disabled')
    
    def _set_status(self, text):
        """Update status bar text. See _log for why this is not update()."""
        if self.shutting_down:
            return
        self.status_text.config(text=text)
        self.root.update_idletasks()
    
    # =========================================================================
    # CONNECTION
    # =========================================================================
    
    def _toggle_connection(self):
        """Connect or disconnect."""
        if self.connection.is_connected:
            self._disconnect()
        else:
            self._connect()
    
    def _disconnect(self):
        """
        Disconnect from Dremio.

        The Connect/Disconnect button stays enabled during execution on
        purpose - it is the user's escape hatch, and the natural thing to reach
        for when a query is taking too long. What used to happen then was a
        race: disconnect nulled the client under the running worker and the
        export died with "'NoneType' object has no attribute 'do_get'" (F-14).

        So disconnecting mid-query is now a deliberate cancellation rather than
        a collision. The user is told what it will cost, the query is cancelled
        through the same path as Stop, and only then does the transport go.
        """
        if self.is_running:
            proceed = messagebox.askyesno(
                "Query Running",
                "A query is still running.\n\n"
                "Disconnecting will cancel it, and no file will be exported.\n\n"
                "Disconnect anyway?"
            )
            if not proceed:
                return
            self._stop_execution()

        self.connection.disconnect()
        self._update_connection_status(False)
        self._log("Disconnected")
        self._set_status("Disconnected")
    
    def _set_connecting_state(self, connecting):
        """
        Put the Connect button into (or out of) its in-flight state.

        Kept separate from the worker so that a worker which never starts can
        still be undone - see _start_worker.
        """
        if connecting:
            self.connect_btn.config(state='disabled', text="Connecting...")
        else:
            self.connect_btn.config(state='normal', text="Connect")

    def _set_executing_state(self, executing):
        """Put the Execute/Stop pair into (or out of) their in-flight state."""
        self.is_running = executing
        if executing:
            self.execute_btn.config(state='disabled')
            self.stop_btn.config(state='normal', bg=COLORS['button_danger'])
        else:
            self.execute_btn.config(state='normal')
            self.stop_btn.config(state='disabled', bg=COLORS['button_secondary'])

    def _start_worker(self, thread, on_failure, what):
        """
        Start a worker thread, undoing the UI state if it will not start.

        Both handlers disable their buttons before starting the worker, and the
        only code that re-enables them lives in the worker's own `finally`. So
        if `start()` raises - RuntimeError: can't start new thread, under thread
        exhaustion - the UI latches permanently with no recovery short of
        restarting the app. This puts the reset somewhere that runs whether or
        not the worker ever exists.
        """
        try:
            thread.start()
            return True
        except Exception as e:
            on_failure()
            self._set_status("Ready")
            self._log(f"ERROR: could not start the {what} thread: {e}")
            messagebox.showerror(
                "Error",
                f"Could not start the {what}:\n\n{e}\n\n"
                f"The application is still usable; try again."
            )
            return False

    def _connect(self):
        """Connect to Dremio."""
        hostname = self.conn_fields['hostname'].get()
        port = self.conn_fields['port'].get()
        username = self.conn_fields['username'].get()
        token = self.conn_fields['token'].get()
        # Snapshotted here with the rest, not read from the worker (F-12).
        use_tls = self.use_tls.get()

        self._log(
            f"Connect requested \u2192 {hostname or '(none)'}:{port or '(none)'} "
            f"as {username or '(none)'} (TLS {'on' if use_tls else 'off'})"
        )

        # Validate, and use what came back rather than what went in: the
        # canonical hostname and port are the ones that were checked (F-24).
        is_valid, error, params = validate_connection_params(
            hostname, port, username, token
        )
        if not is_valid:
            messagebox.showerror("Error", error)
            return

        if params['hostname'] != hostname.strip() or params['port'] != port.strip():
            self._log(f"Connecting to {params['hostname']}:{params['port']}")
        hostname = params['hostname']
        port = params['port']

        # Update UI
        self._set_connecting_state(True)
        self._set_status("Connecting...")

        # Run in background
        thread = threading.Thread(
            target=self._connect_thread,
            args=(hostname, port, username, token, use_tls)
        )
        thread.daemon = True
        self._start_worker(
            thread,
            on_failure=lambda: self._set_connecting_state(False),
            what="connection",
        )
    
    def _connect_thread(self, hostname, port, username, token, use_tls):
        """
        Background connection thread.

        hostname and port arrive canonical - _connect takes them from
        validate_connection_params rather than from the widgets - so there is
        no cleaning step here any more (F-24).

        The `finally` is the point of interest. `token` is a parameter, so it
        lives in this frame, and a frame outlives its call whenever an exception
        carries it away: the traceback holds the frame, the frame holds its
        locals, and the PAT sits in a live object for as long as anything holds
        that exception (F-30). Rebinding it here scrubs the slot on every path,
        including the one that raises - measured, not assumed: without this the
        frame captured in the traceback still reads token='<the PAT>', and with
        it, None.
        """
        t0 = perf_counter()
        try:
            self._ui(lambda: self._log("Starting connection..."))
            self.connection.connect(
                hostname=hostname,
                port=port,
                username=username,
                token=token,
                use_tls=use_tls,
                on_status=lambda msg: self._ui(lambda m=msg: self._log(m))
            )

            elapsed = perf_counter() - t0
            self._ui(lambda e=elapsed: self._log(f"Connection established in {e:.1f}s"))
            # Success
            self._ui(lambda: self._update_connection_status(True, hostname))
            self._ui(lambda: self.connect_btn.config(state='normal'))
            self._ui(lambda: self._set_status(f"Connected to {hostname}"))
            # Queued before the clear below, and the UI queue is FIFO, so the
            # token is still in the widget when this reads it to save.
            self._ui(self._save_current_settings)
            self._ui(self._forget_token_in_form)

        except Exception as e:
            # Bind the text now. Python unbinds `e` when this except block ends,
            # but these lambdas run later, from the event loop - capturing `e`
            # itself makes them raise NameError instead of reporting the failure.
            error_message = str(e)
            self._ui(lambda: self._log(f"ERROR: {error_message}"))
            self._ui(lambda: self.connect_btn.config(state='normal', text="Connect"))
            self._ui(lambda: self._set_status("Connection failed"))
            self._ui(lambda: messagebox.showerror("Connection Failed", error_message))
        finally:
            token = None

    def _forget_token_in_form(self):
        """
        Drop the PAT from the entry once it has been used, unless it is saved.

        Tk thread only.

        The token entry held the PAT for the life of the window: nothing cleared
        it after a successful connect, so it outlived the connection it was for
        and every later one (F-30). What the connection actually needs from then
        on is the derived bearer token, and `disconnect()` already clears that.

        The condition is the honest part. When "Remember token" is ticked the
        PAT is deliberately written to disk, so wiping the widget would be
        theatre - it would still be one file read away, and the user asked for
        that. When it is not ticked, the user has said they do not want it kept,
        and keeping it in a widget for the rest of the session is not honouring
        that.
        """
        if self.remember_token.get():
            return

        self.conn_fields['token'].delete(0, 'end')
        self._log(
            "PAT cleared from the form - 'Remember token' is off. "
            "Re-enter it to reconnect."
        )
    
    # =========================================================================
    # EXECUTION
    # =========================================================================
    
    def _execute_and_export(self):
        """Execute query and export to Excel."""
        if not self.connection.is_connected:
            messagebox.showwarning("Not Connected", "Please connect first!")
            return
        
        query = self.query_text.get('1.0', 'end-1c').strip()
        if not query:
            messagebox.showwarning("No Query", "Please enter a SQL query!")
            return
        
        # Resolve and check the output target before anything expensive runs.
        # Doing it here rather than in the worker is what makes the overwrite
        # prompt possible at all - the worker cannot ask the user anything.
        filepath, error = self._resolve_export_path()
        if error:
            messagebox.showerror("Invalid Filename", error)
            return

        if filepath.exists():
            overwrite = messagebox.askyesno(
                "File Exists",
                f"This file already exists:\n\n{filepath}\n\n"
                f"Exporting will replace it permanently. Overwrite?"
            )
            if not overwrite:
                self._log(f"Export cancelled - {filepath.name} already exists")
                self._set_status("Ready")
                return

        # Add to history
        self.config.add_to_history(query)
        self._update_history_dropdown()

        # Update UI. Clear any cancellation left over from a previous run
        # before the worker can observe it (F-13).
        self.cancel_requested.clear()
        self._set_executing_state(True)

        # Snapshot every widget the worker will need, here on the Tk thread.
        # See _snapshot_export_settings (F-12).
        preview = ' '.join(query.split())
        if len(preview) > 120:
            preview = preview[:117] + '...'
        self._log(f"Execute requested \u2192 {filepath.name}")
        self._log(f"Query: {preview}")

        export_settings = self._snapshot_export_settings(filepath)

        # Run in background
        thread = threading.Thread(
            target=self._execute_thread, args=(query, export_settings)
        )
        thread.daemon = True
        self._start_worker(
            thread,
            on_failure=lambda: self._set_executing_state(False),
            what="query execution",
        )

    def _open_exported_file(self, filepath):
        """
        Show the exported file in the OS file browser, where the platform can.

        Convenience only, and deliberately called OUTSIDE the export's try
        block. os.startfile exists on Windows alone, so when this sat inside the
        try it turned every successful export on Linux and macOS into a reported
        failure: the file was written, os.startfile raised AttributeError, and
        the handler ran instead of the success dialog.

        Failures here are logged as a note. They are not export failures.
        """
        if not hasattr(os, 'startfile'):
            self._ui(lambda: self._log(
                "Note: 'Open after export' is supported on Windows only. "
                "The file was written but not opened."
            ))
            return

        try:
            os.startfile(filepath)
        except Exception as e:
            message = str(e)
            self._ui(lambda: self._log(
                f"Note: the file was exported but could not be opened: {message}"
            ))

    def _execute_thread(self, query, export_settings):
        """Background execution thread."""
        filepath = None
        try:
            def update_progress(msg):
                self._ui(lambda: self.progress_label.config(text=msg))
                self._ui(lambda: self._log(msg))
            
            # Execute query
            q0 = perf_counter()
            self.df = self.connection.execute_query(
                query, on_status=update_progress,
                cancel_event=self.cancel_requested
            )

            # Stop may have been pressed after the last batch arrived but
            # before the export starts. Cancelling then still has to mean no
            # file (F-13).
            if self.cancel_requested.is_set():
                self.df = None
                raise QueryCancelled("Cancelled before the export started")

            rows, cols = len(self.df), len(self.df.columns)
            update_progress(
                f"Retrieved {rows:,} rows × {cols} columns in {perf_counter()-q0:.1f}s"
            )
            
            # Export
            update_progress("Exporting to Excel...")
            x0 = perf_counter()
            filepath = self._export_to_excel(export_settings)
            update_progress(f"Wrote {filepath.name} in {perf_counter()-x0:.1f}s")
            
            update_progress(f"Done! Exported to {filepath.name}")

            # Report any cells Excel could not store in full (F-03). This must
            # happen before the success dialog, so it cannot be missed by a user
            # who dismisses that one and moves on.
            for warning_text in self.last_export_warnings:
                self._ui(lambda w=warning_text: self._log(f"WARNING: {w}"))

            # Control characters Excel cannot carry (F-05). Reported the same
            # way and for the same reason as truncation: the workbook differs
            # from the query result, so saying nothing would be the defect.
            sanitised = self.last_sanitised
            if sanitised:
                total_removed = sum(count for _, _, count in sanitised)
                self._ui(lambda: self._log(
                    f"WARNING: {len(sanitised):,} cell(s) contained control "
                    f"characters Excel cannot store; {total_removed:,} "
                    f"character(s) were removed"
                ))
                for line in self._sanitised_report_lines(sanitised):
                    self._ui(lambda line=line: self._log(f"  {line}"))
                sanitised_text = (
                    f"{len(sanitised):,} cell(s) contained control characters "
                    f"that Excel cannot store - a NUL or similar, usually from "
                    f"a mainframe extract or a fixed-width file.\n\n"
                    f"{total_removed:,} character(s) were removed so the file "
                    f"could be written. Everything else in those cells is "
                    f"unchanged.\n\n"
                    + "\n".join(self._sanitised_report_lines(sanitised))
                )
                self._ui(lambda: messagebox.showwarning(
                    "Control Characters Removed", sanitised_text
                ))

            # Values a spreadsheet would run as a formula, quoted so they stay
            # text (CWE-1236). Reported like sanitisation: the workbook differs
            # from the query result, so saying nothing would be the defect.
            neutralised = self.last_neutralised
            if neutralised:
                self._ui(lambda: self._log(
                    f"WARNING: {len(neutralised):,} cell(s) began with a formula "
                    f"character (= + - @) and were prefixed with an apostrophe so "
                    f"Excel keeps them as text"
                ))
                for line in self._neutralised_report_lines(neutralised):
                    self._ui(lambda line=line: self._log(f"  {line}"))
                neutralised_text = (
                    f"{len(neutralised):,} cell(s) began with a character Excel "
                    f"would treat as a formula (= + - @). Each was prefixed with "
                    f"an apostrophe so it is stored as text, not executed - a "
                    f"value like =HYPERLINK(...) in warehouse data would otherwise "
                    f"run when the workbook is opened.\n\n"
                    + "\n".join(self._neutralised_report_lines(neutralised))
                )
                self._ui(lambda: messagebox.showwarning(
                    "Formula-Like Values Quoted", neutralised_text
                ))

            truncation = self.last_truncation
            if truncation:
                warning_text = truncation['message']
                self._ui(lambda: self._log(
                    f"WARNING: {truncation['count']:,} cell(s) truncated, "
                    f"{truncation['total_lost']:,} characters cut"
                ))
                for line in truncation['lines']:
                    self._ui(lambda line=line: self._log(f"  {line}"))
                if truncation['sidecar']:
                    self._ui(lambda: self._log(
                        f"  full values saved to {truncation['sidecar'].name}"
                    ))
                self._ui(lambda: messagebox.showwarning(
                    "Cells Truncated", warning_text
                ))

            success_text = f"Export complete!\n\n{filepath}\n\n{rows:,} rows exported."
            if truncation:
                success_text += (
                    f"\n\nNote: {truncation['count']:,} cell(s) were truncated - "
                    f"see the warning above."
                )
            if sanitised:
                success_text += (
                    f"\n\nNote: control characters were removed from "
                    f"{len(sanitised):,} cell(s) - see the warning above."
                )
            self._ui(lambda: messagebox.showinfo("Success", success_text))
            export_succeeded = True

        except QueryCancelled as cancelled:
            # Not a failure: the user asked for this. No error dialog, no
            # success dialog, no file - and say plainly that nothing was kept.
            detail = str(cancelled)
            export_succeeded = False
            self.df = None
            self._ui(lambda: self._log(
                f"Execution cancelled - no file was written ({detail})"
            ))
            self._ui(lambda: self.progress_label.config(text="Cancelled"))
            self._ui(lambda: self._set_status("Ready"))

        except Exception as e:
            # Bind the text now - see _connect_thread. `e` is gone by the time
            # the event loop runs these.
            error_message = str(e)
            export_succeeded = False
            self._ui(lambda: self._log(f"ERROR: {error_message}"))
            self._ui(lambda: self.progress_label.config(text="Error"))
            self._ui(lambda: messagebox.showerror("Error", error_message))
        finally:
            # Drop the result frame on every path (F-08). Nothing outside this
            # method reads self.df - the export helpers all run inside
            # _export_to_excel, and the reports above were built from locals
            # before this point - so keeping it only meant the previous result
            # stayed resident for the life of the process. Running a second
            # large query then paid for both frames at once, on top of whatever
            # the export itself needed.
            #
            # The cancel path already cleared it to mean "nothing was kept";
            # this is the same statement made about every path.
            self.df = None
            self._ui(lambda: self._set_executing_state(False))

        # Outside the try on purpose (F-16): the export is already complete and
        # reported by this point, so nothing that happens here may be presented
        # as an export failure.
        if (export_succeeded and filepath is not None
                and export_settings['open_after']):
            self._open_exported_file(filepath)
    
    def _check_row_ceiling(self):
        """
        Refuse a frame with more rows than a worksheet can hold, before writing.

        openpyxl enforces the ceiling at cell-construction time - part-way
        through the write - and the ValueError it raises says only "Row numbers
        must be between 1 and 1048576. Row number supplied was 1048577". It
        names neither Excel nor the query, and by then a partial workbook is
        already at the target path (F-04, F-11).

        The true limit is one row below the sheet ceiling: the header occupies
        sheet row 1, so data starts at row 2.

        Raises:
            ValueError: with a message that says what happened and what to do
                about it.
        """
        rows = len(self.df)
        if rows <= EXCEL_MAX_DATA_ROWS:
            return

        raise ValueError(
            f"The result has {rows:,} rows, which is more than one Excel "
            f"worksheet can hold.\n\n"
            f"A worksheet holds {EXCEL_MAX_SHEET_ROWS:,} rows, and the column "
            f"header uses the first one, leaving {EXCEL_MAX_DATA_ROWS:,} for "
            f"data - {rows - EXCEL_MAX_DATA_ROWS:,} too many here.\n\n"
            f"Add a LIMIT clause to the query, or narrow it with a WHERE "
            f"clause, and run it again. Nothing has been written."
        )

    def _check_column_ceiling(self):
        """
        Refuse a frame with more columns than a worksheet can hold, before
        writing.

        The mirror of _check_row_ceiling for width. A worksheet holds
        EXCEL_MAX_COLUMNS columns; unlike rows, no header line is consumed, so
        the data limit is the sheet limit. openpyxl does not enforce it in
        write_only mode - it appends whatever a row carries, and
        get_column_letter keeps returning references past the real limit - so a
        wider frame yields a workbook Excel refuses to open or silently repairs,
        after the export has already reported success. Checked here, before
        anything is written.

        Raises:
            ValueError: naming the count and what to do about it.
        """
        cols = len(self.df.columns)
        if cols <= EXCEL_MAX_COLUMNS:
            return

        raise ValueError(
            f"The result has {cols:,} columns, which is more than one Excel "
            f"worksheet can hold.\n\n"
            f"A worksheet holds {EXCEL_MAX_COLUMNS:,} columns - "
            f"{cols - EXCEL_MAX_COLUMNS:,} too many here.\n\n"
            f"Select fewer columns in the query and run it again. Nothing has "
            f"been written."
        )

    def _sanitise_illegal_characters(self):
        """
        Replace the control characters openpyxl refuses, and say which cells.

        openpyxl raises IllegalCharacterError part-way through the write, and
        the message - "ab cannot be used in worksheets." - names no column, no
        row and no byte. The offending character is a control code, so it is
        invisible in the dialog and the two strings in that message look
        identical (F-05). There is no way to find the cell from the UI.

        Failing the export is the wrong contract: the user has already paid for
        the query, the transfer and the conversion, and gets nothing back for
        it. Silently stripping is the other wrong contract - that is F-03's
        lesson, and silence is the option off the table. So this replaces and
        reports, as F-03 does.

        Only the affected columns are copied. self.df is left alone, and a
        shallow copy shares the untouched columns rather than duplicating them -
        this runs on frames that F-08 already measures at 33x amplification, so
        a whole-frame copy here would be a real cost rather than a tidiness.

        Column names are sanitised too: a header openpyxl rejects kills the
        export exactly as a cell does.

        Returns:
            tuple: (frame_to_write, affected), where affected is a list of
                (sheet_row, column_name, removed_count) with sheet_row 1-based
                as Excel shows it, and 1 meaning the header itself.
        """
        affected = []
        cleaned_columns = {}

        for position, column in enumerate(self.df.columns):
            # Positional access for the same reason _find_oversized_cells uses
            # it: self.df[column] is a DataFrame, not a Series, when names
            # repeat (F-06).
            series = self.df.iloc[:, position]

            if not (series.dtype == object
                    or pd.api.types.is_string_dtype(series)):
                continue

            as_text = series.astype(str)
            hits = as_text.str.contains(
                ILLEGAL_CHARACTERS_RE, regex=True, na=False
            )
            if not hits.any():
                continue

            for row_position in hits.to_numpy().nonzero()[0]:
                value = as_text.iat[int(row_position)]
                affected.append((
                    int(row_position) + 2,
                    column,
                    len(ILLEGAL_CHARACTERS_RE.findall(value)),
                ))

            # map() over the original values rather than over as_text: an object
            # column can hold non-strings, and writing the str-cast back would
            # turn every NaN into the string 'nan'.
            cleaned_columns[position] = series.map(
                lambda v: ILLEGAL_CHARACTERS_RE.sub(ILLEGAL_CHAR_REPLACEMENT, v)
                if isinstance(v, str) else v
            )

        cleaned_headers = [
            ILLEGAL_CHARACTERS_RE.sub(ILLEGAL_CHAR_REPLACEMENT, str(column))
            for column in self.df.columns
        ]
        headers_changed = cleaned_headers != [str(c) for c in self.df.columns]
        if headers_changed:
            for original, cleaned in zip(self.df.columns, cleaned_headers):
                removed = len(ILLEGAL_CHARACTERS_RE.findall(str(original)))
                if removed:
                    affected.append((1, cleaned, removed))

        if not affected:
            return self.df, []

        frame = self.df.copy(deep=False)
        for position, cleaned in cleaned_columns.items():
            frame.isetitem(position, cleaned)
        if headers_changed:
            frame.columns = cleaned_headers

        return frame, affected

    def _neutralise_formula_cells(self, frame):
        """Quote cells a spreadsheet would run as a formula; say which (CWE-1236).

        A text value beginning '=', '+', '-' or '@' (or a leading TAB/CR a
        spreadsheet trims first) is interpreted as a formula when the workbook is
        opened. Warehouse rows are written by other users, so a cell like
        =HYPERLINK(...) would execute on whoever opens the file. Each such cell is
        prefixed with an apostrophe - Excel's text marker - and the affected cells
        are named, the same replace-and-say contract as F-05. This mirrors the
        dremio_excel skill so both exporters behave identically.

        Runs on the post-sanitise frame. Only object/string columns are touched,
        and only real strings starting with a trigger - numeric columns hold real
        numbers and are left alone. Positional access, like the sibling passes,
        so repeated column names (F-06) stay correct.

        Returns:
            tuple: (frame_to_write, affected), affected a list of
                (sheet_row, column) with sheet_row 1-based as Excel shows it.
        """
        def needs_quoting(value):
            return isinstance(value, str) and value[:1] in FORMULA_TRIGGER_CHARS

        affected = []
        quoted_columns = {}
        for position, column in enumerate(frame.columns):
            series = frame.iloc[:, position]
            if not (series.dtype == object
                    or pd.api.types.is_string_dtype(series)):
                continue
            mask = series.map(needs_quoting)
            if not mask.any():
                continue
            for row_position in mask.to_numpy().nonzero()[0]:
                affected.append((int(row_position) + 2, column))
            quoted_columns[position] = series.map(
                lambda v: (FORMULA_PREFIX + v) if needs_quoting(v) else v
            )

        if not affected:
            return frame, []

        # frame may be self.df on the common path (nothing sanitised), so copy
        # before writing back - the user's result set is not ours to edit.
        out = frame.copy(deep=False)
        for position, quoted in quoted_columns.items():
            out.isetitem(position, quoted)
        return out, affected

    @staticmethod
    def _excel_value(value):
        """
        Convert one pandas value into something openpyxl will accept.

        `df.to_excel` did this for us. Streaming the rows ourselves (F-08) means
        doing it here, and doing it wrong would change exported data silently -
        which makes this the riskiest few lines of that change. Appending raw
        pandas values is not an option: openpyxl raises ValueError on pd.NA,
        NaT and numpy scalars, so the export would fail outright on any nullable
        column.

        The reference for "correct" is not a specification, it is pandas: the
        repro writes the same frame both ways and requires the workbooks to read
        back identical, across text, nullable Int64, float NaN, datetime/NaT,
        bool-with-None, Categorical and Decimal.

        Anything unrecognised becomes str(value), which is what the old autofit
        pass assumed of every value anyway.
        """
        if value is None or value is pd.NaT or value is pd.NA:
            return None

        # Checked before the general pd.isna, which raises on array-likes.
        if isinstance(value, float) and math.isnan(value):
            return None
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass

        if isinstance(value, np.bool_):
            return bool(value)
        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.floating):
            return float(value)
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, np.datetime64):
            return pd.Timestamp(value).to_pydatetime()

        if isinstance(value, (str, int, float, bool, decimal.Decimal,
                              datetime, date, time)):
            return value

        return str(value)

    def _column_widths(self, frame):
        """
        Width for each column, measured before any row is written.

        write_only workbooks accept column dimensions only before the first
        row is appended, so this cannot be folded into the write loop the way
        the old code folded it in after `to_excel`.

        The measurement is per column and transient - one `astype(str)` at a
        time, which the audit measured at +12 MB against the +303 MB the
        workbook itself used to cost.
        """
        widths = []
        for idx, column in enumerate(frame.columns):
            # Positional access: frame[column] returns a DataFrame rather than
            # a Series when two columns share a name - `SELECT a.id, b.id` -
            # and max(Series, int) then raises "truth value of a Series is
            # ambiguous" (F-06).
            series = frame.iloc[:, idx]

            longest = series.astype(str).str.len().max()
            # .max() of an empty Series is nan, and every comparison with nan
            # is False, so max(nan, n) returns nan and the width serialises as
            # width="" - not a valid xsd:double (F-02).
            if pd.isna(longest):
                longest = 0

            width = max(int(longest), len(str(column))) + 2
            widths.append(min(width, 50))
        return widths

    def _find_oversized_cells(self):
        """
        Find cells Excel cannot store in full.

        openpyxl does not raise on an over-length value: it writes the string
        truncated to 32,767 characters. In the write_only mode this export uses
        (F-08) it does not even warn - `Cell.check_string` slices the value and
        returns, and the "Cell contents too long" UserWarning the classic path
        produced came from pandas, which is no longer in the write path.
        Nothing downstream can tell it happened, so the export has to look for
        it before writing.

        Returns:
            list of (sheet_row, column_name, length, full_value), where
            sheet_row is the 1-based row as it appears in Excel (the header
            occupies row 1, so data starts at row 2).
        """
        oversized = []

        for position, column in enumerate(self.df.columns):
            # Positional access on purpose: self.df[column] returns a DataFrame
            # rather than a Series when column names repeat, and this must not
            # add a second failure mode on top of that.
            series = self.df.iloc[:, position]

            # Only text can exceed the limit, and str-casting every numeric
            # column would cost a full transient copy each for nothing.
            if not (series.dtype == object or pd.api.types.is_string_dtype(series)):
                continue

            as_text = series.astype(str)
            lengths = as_text.str.len()
            over = (lengths > EXCEL_MAX_CELL_CHARS).to_numpy().nonzero()[0]

            for row_position in over:
                value = as_text.iat[int(row_position)]
                oversized.append(
                    (int(row_position) + 2, column, len(value), value)
                )

        return oversized

    def _write_truncation_sidecar(self, filepath, oversized):
        """
        Write the full, untruncated values next to the workbook.

        Written incrementally rather than assembled in memory: the whole point
        is that these values are large, and there may be many of them.
        """
        sidecar = filepath.with_name(f"{filepath.stem}.truncated.txt")

        # Atomic temp->replace: a crash mid-write must not leave a partial file
        # of the very values it exists to preserve. Matches the workbook's own
        # no-partial-file contract (F-11) and the dremio_excel skill.
        tmp = sidecar.with_name(sidecar.name + ".tmp")
        with open(tmp, 'w', encoding='utf-8') as f:
            f.write(f"Full values for cells truncated in {filepath.name}\n")
            f.write(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(
                f"\nExcel cannot store more than {EXCEL_MAX_CELL_CHARS:,} "
                f"characters in one cell. The workbook contains the shortened "
                f"values; the complete ones are below.\n"
            )
            f.write(f"\n{len(oversized)} cell(s) affected.\n\n")

            for sheet_row, column, length, value in oversized:
                lost = length - EXCEL_MAX_CELL_CHARS
                f.write("=" * 78 + "\n")
                f.write(
                    f'sheet row {sheet_row}, column "{column}" - '
                    f'{length:,} characters ({lost:,} cut from the workbook)\n'
                )
                f.write("=" * 78 + "\n")
                f.write(value)
                f.write("\n\n")

        os.replace(tmp, sidecar)
        return sidecar

    def _snapshot_export_settings(self, filepath):
        """
        Read every widget the export needs, on the thread that owns them.

        Tk widget access from a worker thread is marshalled back to the Tk
        thread and blocks until it can run; once the window is closed it raises
        outright, stranding the worker (F-12). `_connect` already snapshots its
        inputs this way before starting its thread.

        Returns a plain dict, so nothing downstream holds a widget reference.

        `filepath` is the path the caller already resolved and checked for
        overwrite (F-25); it is passed in rather than resolved again so the
        file that was checked is the file that gets written - re-resolving
        would re-stamp the timestamp and could name a different file.
        """
        return {
            'filepath': filepath,
            'autofit': self.autofit.get(),
            'freeze_header': self.freeze_header.get(),
            # Read here rather than in _execute_thread, which is a worker: the
            # "Open after export" checkbox is the last of the six reads F-12
            # found on the wrong thread.
            'open_after': self.open_after.get(),
            'sheet_name': self._resolve_sheet_name(),
        }

    def _resolve_sheet_name(self):
        """
        The sheet name to write, from config, checked before it is used.

        `output.sheet_name` has been in DEFAULT_CONFIG and written to
        config.json since the beginning, and nothing ever read it - the name was
        the hardcoded literal 'Data'. A user who edited that setting saw no
        effect and no error, which is what F-07 means by the config lying.

        Making it real needs validation in the same move. Nothing had ever
        reached Excel's sheet-name rules precisely because nothing read the
        value, and openpyxl enforces only some of them - see
        utils.validate_sheet_name for which.

        A bad value here does not fail the export. The workbook is fine and the
        user may not know the setting exists; they are told what was wrong and
        what was used instead.
        """
        configured = self.config.get('output', 'sheet_name', DEFAULT_SHEET_NAME)
        is_valid, error, cleaned = validate_sheet_name(configured)

        if is_valid:
            return cleaned

        self._log(
            f"WARNING: the configured sheet name {configured!r} cannot be "
            f"used - {error}. Writing the sheet as {DEFAULT_SHEET_NAME!r} "
            f"instead."
        )
        return DEFAULT_SHEET_NAME

    def _resolve_export_path(self):
        """
        Validate the filename pattern and resolve the full output path.

        Main thread only - it reads widgets.

        Returns:
            tuple: (filepath, error_message). filepath is None on error.
        """
        is_valid, error, pattern = validate_output_filename(self.filename.get())
        if not is_valid:
            return None, error

        filename = generate_timestamp_filename(pattern)
        filepath = Path(self.output_dir.get()) / filename

        # Belt and braces: the pattern was checked for separators, but the
        # substituted result is what actually opens the file.
        try:
            output_dir = Path(self.output_dir.get()).resolve()
            if filepath.resolve().parent != output_dir:
                return None, (
                    f"The filename resolves outside the output folder:\n\n"
                    f"{filepath}\n\nUse a plain filename."
                )
        except OSError as e:
            return None, f"Could not resolve the output path:\n\n{e}"

        return filepath, None

    def _discard_partial_export(self, filepath):
        """
        Remove the half-written workbook left behind by a failed export.

        Best effort by design: this runs while an exception is propagating, and
        failing to delete the remnant must not replace the real error with a
        second one. If it cannot be removed the path is logged, so the user at
        least knows the file is not a real export.
        """
        try:
            if filepath.exists():
                filepath.unlink()
                return True
        except OSError as e:
            message = f"{filepath.name}: {e}"
            self._ui(lambda: self._log(
                f"Note: could not remove the incomplete export {message}"
            ))
        return False

    def _export_to_excel(self, settings=None):
        """
        Export DataFrame to Excel file, streaming the rows.

        The workbook is written with openpyxl's write_only mode rather than
        df.to_excel (F-08). to_excel materialises a Cell object per value, which
        the audit measured at +303 MB of the 539 MB an export cost - more than
        the entire Arrow-to-pandas conversion by an order of magnitude, and the
        reason a 14 MB result set needed ~450 MB to write.

        Measured on the same 100,000 x 10 frame, before and after: the export
        stage itself went from 432 MB to 36 MB (453 to 37 bytes per cell), and
        the whole path from 31x the source Arrow buffers to 7x. What is left is
        mostly the frame itself, which the app has to have. A 1,000,000 x 10
        export projects at 1.0 GB where it projected 4.8 GB.

        Three consequences worth knowing, because they are not obvious:

        - pandas cannot be told to do this. Its openpyxl writer assigns cells by
          coordinate and a WriteOnlyWorksheet has no `.cell`, so the rows are
          appended here.
        - values must be converted before they are appended. to_excel did that;
          _excel_value does it now, and openpyxl raises on pd.NA, NaT and numpy
          scalars, so this is not optional.
        - column widths and freeze_panes must be set before the first row, and
          openpyxl emits no truncation warning in this mode - hence the
          append-time count that replaces it.

        Args:
            settings: dict from _snapshot_export_settings(). Required when
                called from a worker thread; defaults to taking the snapshot
                itself when called on the Tk thread.
        """
        if settings is None:
            if threading.current_thread() is not threading.main_thread():
                raise RuntimeError(
                    "_export_to_excel requires an export-settings snapshot when "
                    "called off the main thread. Call _snapshot_export_settings() "
                    "on the Tk thread and pass the result (F-12)."
                )
            filepath, error = self._resolve_export_path()
            if error:
                raise ValueError(error)
            settings = self._snapshot_export_settings(filepath)

        self.last_truncation = None
        self.last_export_warnings = []
        self.last_sanitised = []
        self.last_neutralised = []

        # Before anything is created on disk. openpyxl would raise this
        # part-way through the write instead, leaving a partial workbook behind
        # and a message that mentions neither Excel nor the query (F-04).
        self._check_row_ceiling()
        self._check_column_ceiling()

        filepath = settings['filepath']
        filepath.parent.mkdir(parents=True, exist_ok=True)

        # Look for over-length cells before writing, so the values are still
        # available to preserve.
        oversized = self._find_oversized_cells()

        # Replace the control bytes openpyxl refuses, and remember which cells
        # they came from (F-05). frame is self.df itself when there is nothing
        # to change, so the common path copies nothing.
        frame, self.last_sanitised = self._sanitise_illegal_characters()

        # Quote cells a spreadsheet would run as a formula (CWE-1236), and
        # remember which. Runs on the post-sanitise frame; copies only if it has
        # to change something. Matches the dremio_excel skill's neutralisation.
        frame, self.last_neutralised = self._neutralise_formula_cells(frame)

        # Export.
        #
        # The write is still wrapped in catch_warnings, but no longer for
        # truncation: openpyxl raises no warning for an over-length cell, and
        # the one the old path captured came from pandas, which write_only mode
        # removes from the picture. What remains worth catching is everything
        # else openpyxl may say during a write - it says it on stderr, which a
        # windowed PyInstaller build does not have, so left alone it goes
        # nowhere. The truncation cross-check is now the append-time count
        # below.
        #
        # catch_warnings manipulates global state and is not thread-safe. That
        # is acceptable here only because execute_btn is disabled for the
        # duration, so at most one export runs at a time.
        # `frame`, not self.df: it is the same object unless
        # _sanitise_illegal_characters had to replace something, in which case
        # it is the version openpyxl will accept (F-05). The sheet name was
        # checked by _resolve_sheet_name before it got here (F-07); the fallback
        # covers _export_to_excel being called without a snapshot.
        sheet_name = settings.get('sheet_name') or DEFAULT_SHEET_NAME

        widths = self._column_widths(frame) if settings['autofit'] else None
        # A count, not a list. This is the one thing in the write loop that
        # grows with the data, and a whole over-limit column would otherwise
        # build a message per cell that nothing ever reads - which is the
        # failure mode this change exists to remove.
        written_over_limit = 0

        try:
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")

                # write_only, rather than df.to_excel (F-08). to_excel builds an
                # in-memory Cell object per value, which measured +303 MB of the
                # 539 MB an export cost - the dominant term by an order of
                # magnitude over the whole Arrow-to-pandas conversion. Streaming
                # the rows instead took the export stage from 432 MB to 36 MB on
                # the same frame.
                #
                # pandas cannot be asked to do this. Its openpyxl writer sets
                # cells by coordinate, and a WriteOnlyWorksheet has no .cell -
                # so the rows are appended here directly.
                workbook = Workbook(write_only=True)
                ws = workbook.create_sheet(title=sheet_name)

                # Both of these must be set BEFORE the first row is appended;
                # a write_only sheet has already streamed past them afterwards.
                if widths is not None:
                    for position, width in enumerate(widths, start=1):
                        # openpyxl's own converter. The previous arithmetic
                        # walked off the end of the alphabet at index 52,
                        # emitting 'A[', 'A\', 'A]' (F-01).
                        ws.column_dimensions[get_column_letter(position)].width = width
                if settings['freeze_header']:
                    ws.freeze_panes = 'A2'

                ws.append([str(column) for column in frame.columns])

                # itertuples rather than iterrows: iterrows builds a Series per
                # row, which would put back a per-row allocation this change
                # exists to remove.
                for row in frame.itertuples(index=False, name=None):
                    values = []
                    for value in row:
                        value = self._excel_value(value)
                        # Counted as it is handed over, which is what replaces
                        # openpyxl's truncation warning - it does not emit one
                        # here, but still truncates at the limit. See the note
                        # in _export_to_excel's docstring.
                        if (isinstance(value, str)
                                and len(value) > EXCEL_MAX_CELL_CHARS):
                            written_over_limit += 1
                        values.append(value)
                    ws.append(values)

                workbook.save(filepath)
        except Exception:
            # A failure mid-write can still leave a file at the target path
            # (F-11). It used to arrive via pd.ExcelWriter.__exit__, which calls
            # close() -> save() without inspecting the exception state; now it
            # arrives from workbook.save() itself failing part-way through the
            # zip. Either way, what is left there is indistinguishable from a
            # real export - same filename pattern, same folder - and is either a
            # metadata-only zip Excel offers to repair, or something that opens
            # cleanly with the right headers and no rows.
            #
            # Nothing is lost by removing it. save() opened the path in write
            # mode, which truncated any previous file before the failure, so the
            # old contents were already gone.
            self._discard_partial_export(filepath)
            raise

        # `written_over_limit` is the independent count of what was actually
        # truncated, and it replaces a check that would otherwise have gone
        # quietly wrong.
        #
        # This used to reconcile the pre-write scan against openpyxl's own "too
        # long" UserWarning. That warning does not exist on this path: openpyxl
        # truncates silently, and the message the old code matched on came from
        # pandas' writer. Keeping it would have reported zero truncations
        # against every scan that found some - so every truncating export would
        # have told the user the sidecar might be incomplete when it was fine.
        #
        # Counting the values as they are handed over is a better cross-check
        # than the warning ever was, and it is only available because this code
        # now owns the write loop: the scan reads the frame beforehand, this
        # observes what was actually written, and the two agreeing is what makes
        # the sidecar trustworthy.
        self.last_export_warnings = [
            f"{w.category.__name__}: {w.message}" for w in caught
        ]

        if oversized or written_over_limit:
            self.last_truncation = self._record_truncation(
                filepath, oversized, written_over_limit
            )

        return filepath

    def _sanitised_report_lines(self, sanitised):
        """
        Name the cells whose control characters were removed (F-05).

        Bounded like the truncation report: naming every cell in a dialog is
        unreadable when a whole column is affected, which is the usual case for
        a padded mainframe extract. Row 1 is the header, so it is labelled as
        such rather than shown as a row number.
        """
        listed = sanitised[:SANITISED_REPORT_LIMIT]
        lines = [
            (f'column header "{col}" - {count:,} removed' if row == 1 else
             f'sheet row {row}, column "{col}" - {count:,} removed')
            for row, col, count in listed
        ]
        if len(sanitised) > len(listed):
            lines.append(
                f"...and {len(sanitised) - len(listed):,} more"
            )
        return lines

    def _neutralised_report_lines(self, neutralised):
        """
        Name the cells that were quoted to stop them running as a formula.

        Bounded like the sanitised and truncation reports - a whole affected
        column is common, and naming every cell in a dialog is unreadable.
        """
        listed = neutralised[:SANITISED_REPORT_LIMIT]
        lines = [f'sheet row {row}, column "{col}"' for row, col in listed]
        if len(neutralised) > len(listed):
            lines.append(f"...and {len(neutralised) - len(listed):,} more")
        return lines

    def _record_truncation(self, filepath, oversized, written_over_limit):
        """
        Preserve the full values and build the report shown to the user.

        A failure to write the sidecar must not fail an export that otherwise
        succeeded - the workbook is already on disk and is still useful. It does
        change the message, which then has to say the values were lost.

        Args:
            oversized: what the pre-write scan of the frame found - the only
                source that can name a column and preserve a value.
            written_over_limit: how many over-length values the write loop
                actually handed to openpyxl. The two are reconciled below.
        """
        total_lost = sum(length - EXCEL_MAX_CELL_CHARS
                         for _, _, length, _ in oversized)

        sidecar = None
        sidecar_error = None
        try:
            sidecar = self._write_truncation_sidecar(filepath, oversized)
        except Exception as e:
            sidecar_error = str(e)

        listed = oversized[:TRUNCATION_REPORT_LIMIT]
        lines = [
            f'sheet row {row}, column "{col}" - {length:,} chars '
            f'({length - EXCEL_MAX_CELL_CHARS:,} cut)'
            for row, col, length, _ in listed
        ]
        if len(oversized) > len(listed):
            lines.append(f"...and {len(oversized) - len(listed):,} more "
                         f"(all of them are in the sidecar file)")

        # Reconcile the pre-write scan against what the write loop actually
        # handed to openpyxl. They should agree; if they do not, the scan missed
        # a cell and the sidecar is incomplete, which the user needs to be told
        # rather than left to discover.
        discrepancy = None
        if written_over_limit != len(oversized):
            discrepancy = (
                f"{written_over_limit:,} over-length value(s) were written but "
                f"the pre-write scan found {len(oversized):,}. The sidecar may be "
                f"incomplete - treat the workbook as lossy and re-run after "
                f"shortening the column in SQL."
            )
            lines.append(f"WARNING: {discrepancy}")

        count = max(len(oversized), written_over_limit)
        headline = (
            f"{count:,} cell(s) exceeded Excel's "
            f"{EXCEL_MAX_CELL_CHARS:,}-character limit and were truncated in the "
            f"workbook ({total_lost:,} characters cut)."
        )

        if sidecar:
            resolution = f"The full values were saved to:\n{sidecar.name}"
        else:
            resolution = (
                f"The full values could NOT be saved - writing the sidecar file "
                f"failed: {sidecar_error}\nThose characters are lost from this "
                f"export. Re-run after shortening the column in SQL."
            )

        return {
            'count': count,
            'scanned': len(oversized),
            'warned': written_over_limit,
            'discrepancy': discrepancy,
            'total_lost': total_lost,
            'sidecar': sidecar,
            'sidecar_error': sidecar_error,
            'lines': lines,
            'message': f"{headline}\n\n" + "\n".join(lines) + f"\n\n{resolution}",
        }
    
    def _stop_execution(self):
        """
        Ask the running query to stop, and mean it.

        This used to set self.is_running = False and log "Execution cancelled" -
        a statement nothing acted on and nothing made true. The query ran to
        completion, the file was written, and the success dialog appeared,
        while the log said the run had been cancelled (F-13).

        Two things now happen instead. The flag is an Event the worker actually
        reads between record batches, and connection.cancel_query() interrupts
        the read the worker is currently blocked in, so Stop takes effect at
        once rather than at the end of the stream.

        is_running is left to the worker's own finally, via
        _set_executing_state: the run is not over until the worker says so, and
        claiming otherwise here is what made the original message false.
        """
        if not self.is_running:
            return

        self.cancel_requested.set()
        self.stop_btn.config(state='disabled')
        self._set_status("Cancelling...")
        self._log("Cancelling - asking the server to stop sending data...")

        if not self.connection.cancel_query():
            # Nothing was mid-stream: either the query has not reached the read
            # yet, or it is already past it. The Event covers the first case -
            # the worker checks it before opening the stream and again before
            # exporting. It does not cover the second: a write already in
            # progress runs to completion rather than leaving a half-file
            # behind, so say that rather than promising otherwise.
            self._log(
                "No data transfer in progress - the query has not started "
                "streaming yet, or the export has already begun and will finish"
            )
