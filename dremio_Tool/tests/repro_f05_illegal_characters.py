"""
================================================================================
F-05 - Characters openpyxl rejects abort the export                  (Medium)
================================================================================
A control byte anywhere in the result killed the export. openpyxl raises
IllegalCharacterError part-way through the write, and the message it produces is

    ab cannot be used in worksheets.

which names no column, no row and no byte. The offending character is a control
code, so it is invisible in the dialog: the two strings in that message look
identical to the user. There was no way to find the bad cell from the UI - and
per F-33 the dialog usually did not appear at all, arriving 1 time in 5.

Column names had the same problem: a rejected header kills the export exactly as
a cell does.

The contract chosen, and why
----------------------------
Three options, two of them wrong:

  - fail the export. The user has already paid for the query, the transfer and
    the conversion, and gets nothing back for a byte Excel cannot carry.
  - strip silently. That is F-03's lesson exactly, and silence is the one option
    off the table: the workbook would differ from the query result with nothing
    saying so.
  - replace, and say which cells. This is what F-03 does for over-length values,
    so the export path now handles both the same way.

Control bytes reach real result sets through mainframe extracts, fixed-width
files and CHAR(n) padding, so this is ordinary data rather than an edge case.

What is checked
---------------
That the export succeeds, that the workbook reads back with the control bytes
gone and everything else intact, that self.df is NOT mutated, and that the user
is told - the last being the part that separates this fix from the silent strip.
================================================================================
"""

REQUIRES_DISPLAY = True

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import harness as h

# The audit's own table: which bytes openpyxl refuses and which it permits.
CHARACTER_CASES = [
    ("\x00", "NUL", False),
    ("\x07", "BEL", False),
    ("\x0b", "VT", False),
    ("\x1f", "US", False),
    ("\t", "TAB", True),
    ("\n", "LF", True),
    ("\U0001F600", "emoji (non-BMP)", True),
]


def what_openpyxl_refuses():
    """Re-measure the boundary rather than trusting the table."""
    h.step("Which characters openpyxl actually refuses")
    rows = []
    for char, name, legal in CHARACTER_CASES:
        rejected = bool(ILLEGAL_CHARACTERS_RE.search(f"a{char}b"))
        agrees = rejected != legal
        rows.append([name, repr(char), "permitted" if legal else "refused",
                     "ok" if agrees else "DISAGREES"])
    h.table(["character", "", "expected", ""], rows)
    return all(r[3] == "ok" for r in rows)


def export_with_control_characters(tmp):
    """
    Drive the real export over a frame carrying control bytes in cells and in a
    column name, and read the workbook back.
    """
    h.step("The real export, over data Excel cannot carry verbatim")

    frame = pd.DataFrame({
        "clean": ["ordinary", "values"],
        "padded": ["a\x00b", "c\x07d"],
        "kept": ["tab\there", "line\nbreak"],
    })
    # A header with a control byte kills the export the same way a cell does.
    frame.columns = ["clean", "pad\x1fded", "kept"]
    original = frame.copy(deep=True)

    path = tmp / "control_chars.xlsx"
    state = {}
    with h.tk_app(output_dir=tmp, filename="control_chars.xlsx",
                  autofit=True) as app:
        app.df = frame
        try:
            app._export_to_excel()
            state["exported"] = True
            state["error"] = ""
        except Exception as e:
            state["exported"] = False
            state["error"] = f"{type(e).__name__}: {str(e)[:80]}"
        # getattr rather than attribute access: a build without the fix has
        # neither, and "nothing was reported" is the finding rather than a
        # reason for this script to die.
        state["reported"] = list(getattr(app, "last_sanitised", []))
        reporter = getattr(app, "_sanitised_report_lines", None)
        state["lines"] = (reporter(state["reported"])
                          if reporter and state["reported"] else [])

    # self.df must be untouched: the user's result set is not the workbook's to
    # edit, and _execute_thread reads its shape afterwards.
    state["df_unmutated"] = frame.equals(original) and list(frame.columns) == list(original.columns)

    if state["exported"] and path.exists():
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]   # first sheet, whatever it is named
        state["headers"] = [c.value for c in ws[1]]
        state["cells"] = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    else:
        state["headers"] = []
        state["cells"] = []

    h.table(["check", "result"], [
        ["export completed", str(state["exported"])
         + (f" ({state['error']})" if state["error"] else "")],
        ["headers written", repr(state["headers"])],
        ["data written", repr(state["cells"])],
        ["cells reported to the user", str(len(state["reported"]))],
        ["self.df left unmutated", str(state["df_unmutated"])],
    ])
    for line in state["lines"]:
        h.detail("reported", line)
    return state


def main():
    h.require_display()
    h.banner("F-05", "Control characters, and whether the export survives them")

    boundary_agrees = what_openpyxl_refuses()

    with h.isolated_home(), h.temp_dir() as tmp:
        state = export_with_control_characters(tmp)

    flat = [value for row in state["cells"] for value in row]
    written = state["headers"] + [v for v in flat if isinstance(v, str)]

    control_bytes_left = [v for v in written if ILLEGAL_CHARACTERS_RE.search(v)]
    # TAB and LF are legal and carry meaning - stripping them would be a
    # different data-loss bug wearing this fix's clothes.
    kept_legal = any("\t" in v for v in written) and any("\n" in v for v in written)
    # The other columns must arrive untouched.
    kept_clean = "ordinary" in written and "values" in written
    reported = len(state["reported"]) > 0

    h.step("Contract check")
    h.detail("openpyxl's boundary is as the audit recorded", boundary_agrees)
    h.detail("export completed", state["exported"])
    h.detail("control bytes remaining in the workbook",
             control_bytes_left or "none")
    h.detail("TAB and LF preserved", kept_legal)
    h.detail("untouched columns intact", kept_clean)
    h.detail("affected cells reported to the user", len(state["reported"]))
    h.detail("self.df left unmutated", state["df_unmutated"])

    if (state["exported"] and not control_bytes_left and kept_legal
            and kept_clean and reported and state["df_unmutated"]):
        h.verdict("F-05", h.NOT_REPRODUCIBLE,
                  f"control characters no longer abort the export: the bytes "
                  f"openpyxl refuses are removed from both cells and column "
                  f"names, TAB and LF - which Excel permits and which carry "
                  f"meaning - are preserved, every other value arrives "
                  f"unchanged, and the {len(state['reported'])} affected cell(s) "
                  f"are named to the user in the log and a dialog rather than "
                  f"changed silently. self.df itself is not mutated")
    elif not state["exported"]:
        h.verdict("F-05", h.CONFIRMED,
                  f"the export still dies on a control character: "
                  f"{state['error']}")
    else:
        h.verdict("F-05", h.CONFIRMED,
                  f"exported, but the contract is not met: "
                  f"control_bytes_left={control_bytes_left} "
                  f"tab_lf_preserved={kept_legal} clean_columns={kept_clean} "
                  f"reported={reported} df_unmutated={state['df_unmutated']}")


if __name__ == "__main__":
    main()
