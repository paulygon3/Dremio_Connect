"""
================================================================================
F-01 - Column-letter formula is wrong for every index >= 52   (High)
F-02 - Auto-fit width is nan on a zero-row result set          (Medium)
================================================================================
AUDIT.md tagged both EXECUTED, proven by the original tests/repro_column_bug.py,
which reimplemented the formula standalone. This version supersedes it: it
drives the real DremioExporter._export_to_excel (app.py:745-774) so that a
Stage 1 fix flips the verdict on its own instead of leaving a stale copy of the
buggy arithmetic passing forever.

F-01 mechanism: app.py:767 computes
    chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
whose else-branch can only ever emit A?. Correct to AZ (idx 51); at idx 52 it
walks off the end of the alphabet into '[', '\\', ']'.

F-02 mechanism: app.py:763-766 takes .max() of an empty Series, which is nan,
and every comparison with nan is False, so max(nan, 1) returns nan.
================================================================================
"""

REQUIRES_DISPLAY = True

import zipfile

import pandas as pd

import harness as h


def check_formula_against_openpyxl():
    """The arithmetic, isolated. Cheap, and pins the exact divergence point."""
    from openpyxl.utils import get_column_letter

    h.step("F-01 step 1: app's column-letter expression vs openpyxl's own")

    def app_formula(idx):
        # transcribed from app.py:767 for the boundary scan only
        return chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"

    first_bad = None
    rows = []
    for idx in (0, 25, 26, 51, 52, 53, 57, 83, 701):
        got = app_formula(idx)
        want = get_column_letter(idx + 1)
        ok = got == want
        if not ok and first_bad is None:
            first_bad = idx
        rows.append([idx, repr(got), want, "ok" if ok else "WRONG"])
    h.table(["idx", "app formula", "correct", ""], rows)

    wrong = sum(1 for i in range(1000) if app_formula(i) != get_column_letter(i + 1))
    h.detail("first divergence at idx", first_bad)
    h.detail("wrong over idx 0..999", f"{wrong} of 1000")
    return first_bad


def real_export_wide(tmp):
    """Drive the real _export_to_excel with 60 columns."""
    h.step("F-01 step 2: real _export_to_excel on a 60-column frame")

    df = pd.DataFrame({f"c{i}": [i] for i in range(60)})
    with h.tk_app(output_dir=tmp, filename="wide.xlsx", autofit=True) as app:
        app.df = df
        h.detail("autofit", app.autofit.get())
        h.detail("columns", len(df.columns))
        try:
            path = app._export_to_excel()
            h.detail("result", f"SAVED {path.name}")
            return None, tmp / "wide.xlsx"
        except Exception as e:
            h.detail("raised", f"{type(e).__name__}: {e}")
            return e, tmp / "wide.xlsx"


def real_export_empty(tmp):
    """Drive the real _export_to_excel with zero rows, then read the raw XML."""
    h.step("F-02: real _export_to_excel on a zero-row frame")

    df = pd.DataFrame({"a": pd.Series([], dtype="str")})
    with h.tk_app(output_dir=tmp, filename="empty.xlsx", autofit=True) as app:
        app.df = df
        # Reproduce the width computation the export performs, to show the nan.
        computed = max(df["a"].astype(str).map(len).max(), len("a")) + 2
        h.detail("max(empty.max(), len(col)) + 2", repr(computed))
        try:
            path = app._export_to_excel()
            h.detail("result", f"SAVED {path.name}")
        except Exception as e:
            h.detail("raised", f"{type(e).__name__}: {e}")
            return None

    xml_path = tmp / "empty.xlsx"
    with zipfile.ZipFile(xml_path) as z:
        sheet = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    start = sheet.find("<cols>")
    end = sheet.find("</cols>")
    cols_xml = sheet[start:end + 7] if start != -1 else "(no <cols> element)"
    h.detail("written <cols>", cols_xml)
    return cols_xml


def main():
    h.require_display()
    h.banner("F-01 / F-02", "Column-letter arithmetic and zero-row auto-fit width")

    with h.isolated_home(), h.temp_dir() as tmp:
        first_bad = check_formula_against_openpyxl()
        exc, _ = real_export_wide(tmp)
        cols_xml = real_export_empty(tmp)

    # --- F-01 verdict ---
    if exc is not None and isinstance(exc, ValueError):
        h.verdict("F-01", h.CONFIRMED,
                  f"real _export_to_excel on 60 cols raised {type(exc).__name__}: {exc}; "
                  f"formula first diverges at idx {first_bad}")
    elif exc is not None:
        h.verdict("F-01", h.CONFIRMED,
                  f"export failed with {type(exc).__name__}: {exc} (expected ValueError)")
    else:
        h.verdict("F-01", h.NOT_REPRODUCIBLE,
                  "60-column export completed without error - column-letter path appears fixed")

    # --- F-02 verdict ---
    if cols_xml is None:
        h.verdict("F-02", h.NOT_REPRODUCIBLE, "zero-row export raised instead of writing a file")
    elif 'width=""' in cols_xml:
        h.verdict("F-02", h.CONFIRMED,
                  'zero-row export writes width="" - not a valid xsd:double')
    elif "nan" in cols_xml.lower():
        h.verdict("F-02", h.CONFIRMED, f"zero-row export writes a nan width: {cols_xml}")
    else:
        h.verdict("F-02", h.NOT_REPRODUCIBLE, f"width looks well-formed: {cols_xml}")


if __name__ == "__main__":
    main()
