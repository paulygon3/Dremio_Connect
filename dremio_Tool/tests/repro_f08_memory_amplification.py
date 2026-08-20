"""
================================================================================
F-08 - Peak-memory amplification between read_all() and the file   (High)
================================================================================
AUDIT.md measured 32x amplification with ru_maxrss and flagged its own caveat:
ru_maxrss is a monotonic high-water mark, so each delta is "how much the peak
rose during this stage", not "how much was resident at that moment". Stage 0
resolved that caveat - the copies are concurrent - and identified the dominant
term: the openpyxl Workbook, one Cell object per value, +303 MB of a 539 MB
export.

This script now has to answer three questions, because the Tier 3 fix (streaming
the rows through a write_only workbook, and releasing the frame afterwards) can
fail in three different ways:

  1. is the memory actually gone?    - the stage table, as before, plus the
                                       cost of the EXPORT STAGE alone, which is
                                       what the fix changes and what the whole
                                       finding is about
  2. is the data still correct?      - a write path this app owns can no longer
                                       lean on to_excel to convert values, so
                                       the same frame is written both ways and
                                       the two workbooks must read back
                                       identical, cell for cell and type for
                                       type
  3. is the frame released?          - the second half of the finding: nothing
                                       cleared self.df, so the previous result
                                       stayed resident and a second large query
                                       paid for both

Question 2 is not optional politeness. It is the risk the fix introduces: a
memory win that silently changed exported values would be far worse than the bug
it replaced, and nothing else in the suite would catch it.

Every stage runs real app code: read_all() over Arrow Flight, the real
connection._arrow_to_pandas(), the real DremioExporter._export_to_excel(), and -
for question 3 - the real Execute handler on the real worker thread. The one
instrumentation point is a wrapper around openpyxl's Workbook.save so workbook
construction and XML/zip serialisation can be told apart; the real save still
runs.

The Flight server runs in a SEPARATE PROCESS. An in-process server holds its own
full copy of the served table inside the process being measured, which inflates
the baseline and every residency figure by the size of the data.

This is the slowest script in the suite (1,000,000 cells through openpyxl).
================================================================================
"""

REQUIRES_DISPLAY = True

import ast
import decimal
import gc
import threading
import warnings

import pandas as pd
import pyarrow as pa

import harness as h

ROWS = 100_000
COLS = 10

# What the export stage may cost, per cell, before this counts as unfixed.
#
# The two states are nowhere near each other, so the threshold does not need to
# be delicate. Measured on this frame: ~420 bytes per cell with df.to_excel
# (a Cell object per value, plus the XML/zip buffers on save), ~40 bytes per
# cell streaming. Anything approaching the old figure means the workbook is
# being built in memory again.
EXPORT_BUDGET_BYTES_PER_CELL = 100


class Sampler:
    """Records (label, peak_mb, current_mb) at each stage."""

    def __init__(self):
        self.rows = []

    def take(self, label):
        gc.collect()
        self.rows.append((label, h.peak_rss_mb(), h.current_rss_mb()))
        return self.rows[-1]

    def add(self, label, peak, current):
        self.rows.append((label, peak, current))

    def find(self, label):
        for row in self.rows:
            if row[0] == label:
                return row
        return None

    def report(self):
        rows = []
        prev = None
        for label, peak, current in self.rows:
            delta = "-" if prev is None else f"+{peak - prev:.0f}"
            rows.append([label, f"{peak:.0f}", delta,
                         f"{current:.0f}" if current is not None else "n/a"])
            prev = peak
        h.table(["stage", "peak RSS MB", "d peak", "current RSS MB"], rows)


def df_clearing_sites():
    """
    Find every `self.df = None` in app.py with its enclosing function.

    A plain grep is misleading here: __init__ sets self.df = None at app.py:96,
    which is initialisation, not release. Only an assignment somewhere else is
    evidence that the frame is ever let go.
    """
    tree = ast.parse(h.source_text("app.py"))
    sites = []
    for func in [n for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)]:
        for node in ast.walk(func):
            if not isinstance(node, ast.Assign):
                continue
            for target in node.targets:
                if (isinstance(target, ast.Attribute) and target.attr == "df"
                        and isinstance(target.value, ast.Name)
                        and target.value.id == "self"
                        and isinstance(node.value, ast.Constant)
                        and node.value.value is None):
                    sites.append((func.name, node.lineno))
    return sites


def self_destruct_probe(label, table):
    """
    After conversion, is copy 1 still holding its Arrow buffers?

    connection.py:372 passes self_destruct=True, whose purpose is to free Arrow
    buffers incrementally as they are converted. AUDIT.md reasons it cannot work
    because copy 1 stays bound to the `table` local in execute_query's frame.

    This is measurement, not a pass/fail check, and the fix does not touch it:
    Stage 0 established that copy 1 pins single-digit MB, so it is not where the
    memory goes. It is re-measured here so that stays on the record.

    Measured with pa.total_allocated_bytes(), NOT with RSS. RSS is the wrong
    instrument for this question: glibc and CPython do not return freed pages to
    the OS promptly, so RSS stays flat whether or not the buffers were released,
    and an RSS-based probe reports "0 MB freed" in every case regardless of the
    truth. Arrow's memory pool accounts for its own buffers exactly.
    """
    h.add_src_to_path()
    from connection import DremioConnection

    mb = 1024.0 * 1024.0
    conn = DremioConnection()
    gc.collect()

    held_before = pa.total_allocated_bytes() / mb
    df = conn._arrow_to_pandas(table)
    gc.collect()
    held_after_conv = pa.total_allocated_bytes() / mb
    del table
    gc.collect()
    held_after_release = pa.total_allocated_bytes() / mb

    h.detail(f"{label}", "")
    h.detail("  Arrow bytes held, table alive, before conversion",
             f"{held_before:.1f} MB")
    h.detail("  Arrow bytes held, DataFrame built, table local STILL BOUND",
             f"{held_after_conv:.1f} MB")
    h.detail("  Arrow bytes held after releasing the table local",
             f"{held_after_release:.1f} MB")
    retained = held_after_conv - held_after_release
    h.detail("  => still pinned by copy 1 during conversion", f"{retained:.1f} MB")

    del df
    gc.collect()
    return retained


def fidelity_frame():
    """
    One frame covering every value shape the conversion has to get right.

    Each column is here because it breaks something naive. Nullable Int64, NaT
    and numpy scalars make openpyxl raise outright, so an export that appended
    raw pandas values would fail on any of them rather than corrupt anything.
    The dangerous ones are the columns that would convert *quietly* to the wrong
    thing: NaN reaching the sheet as the string "nan", a Categorical as its
    integer code, a Decimal as a repr, a bool column holding None.
    """
    return pd.DataFrame({
        "text": ["a", "", None, "unicode: ünïcode"],
        "nullable_int": pd.array([1, 2, None, 4], dtype="Int64"),
        "plain_int": [10, 20, 30, 40],
        "float_with_nan": [1.5, float("nan"), 3.25, -0.5],
        "datetime_with_nat": pd.to_datetime([
            "2024-01-02 03:04:05", None,
            "1999-12-31 00:00:00", "2030-06-15 12:00:00"]),
        "bool_with_none": [True, False, None, True],
        "categorical": pd.Categorical(["x", "y", None, "x"]),
        "decimal": [decimal.Decimal("1.25"), decimal.Decimal("0"), None,
                    decimal.Decimal("-3.5")],
    })


def check_fidelity(tmp):
    """
    The app's export against pandas' own, on the same frame.

    pandas IS the specification here. There is no document that says what a
    pd.NA should look like in a worksheet; what there is, is the behaviour this
    app shipped with, and the fix has to preserve it exactly. So the reference
    workbook is written by the code path being replaced - df.to_excel - and both
    are read back and compared value by value AND type by type.

    Types matter as much as values: a datetime written as the string
    "2024-01-02 03:04:05" reads back equal to nothing and sorts as text in
    Excel, and 1.25 written as "1.25" cannot be summed. Comparing only values
    would let both through.
    """
    from openpyxl import load_workbook

    frame = fidelity_frame()
    h.step("Data fidelity: the same frame, written both ways")
    h.detail("columns", ", ".join(
        f"{name} ({dtype})" for name, dtype in frame.dtypes.astype(str).items()))

    reference = tmp / "reference_to_excel.xlsx"
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        with pd.ExcelWriter(reference, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, sheet_name="Data")

    with h.tk_app(output_dir=tmp, filename="app_export.xlsx",
                  autofit=True) as app:
        app.df = frame
        produced = app._export_to_excel()

    left = load_workbook(reference)["Data"]
    produced_book = load_workbook(produced)
    # Whatever the app named its sheet - that is F-07's business, not this
    # check's.
    right = produced_book[produced_book.sheetnames[0]]

    same_shape = (left.max_row == right.max_row
                  and left.max_column == right.max_column)
    h.detail("reference (df.to_excel) grid",
             f"{left.max_row} x {left.max_column}")
    h.detail("app export grid", f"{right.max_row} x {right.max_column}")

    diffs = []
    if same_shape:
        for row in range(1, left.max_row + 1):
            for col in range(1, left.max_column + 1):
                expected = left.cell(row, col).value
                actual = right.cell(row, col).value
                if expected != actual or type(expected) is not type(actual):
                    diffs.append(
                        [f"r{row}c{col}",
                         f"{expected!r} ({type(expected).__name__})",
                         f"{actual!r} ({type(actual).__name__})"])

    if diffs:
        h.table(["cell", "df.to_excel wrote", "the app wrote"], diffs[:20])
        if len(diffs) > 20:
            h.note(f"...and {len(diffs) - 20} more")
    else:
        h.detail("cells compared", left.max_row * left.max_column)
        h.detail("differences in value or type", 0)

    h.detail("warnings raised by the reference write",
             [str(w.message)[:60] for w in caught] or "none")

    return same_shape and not diffs


def check_frame_released(tmp):
    """
    Does a completed run let go of the result?

    This has to go through the real Execute handler and the real worker thread,
    because the release belongs to _execute_thread - calling _export_to_excel
    directly would leave self.df set no matter what the fix does, and reporting
    that as "never released" would be an artifact of how the test called it.

    A StubConnection stands in for the server: the question is what the app does
    with the frame after the export, which does not depend on where it came
    from.
    """
    h.step("Is the result frame released when the run finishes?")

    frame = pd.DataFrame({f"c{i}": [f"v_{i}_{r}" for r in range(500)]
                          for i in range(4)})
    observed = {}

    with h.tk_app(output_dir=tmp, filename="released.xlsx") as app:
        app.connection = h.StubConnection(df=frame)
        app.query_text.delete("1.0", "end")
        app.query_text.insert("1.0", "SELECT * FROM t")

        def scenario():
            # The success dialog is modal; recorded rather than shown, or the
            # mainloop this runs inside would never come back.
            with h.captured_dialogs() as dialogs:
                before = set(threading.enumerate())
                app._execute_and_export()
                worker = h.new_threads_from(before)
                h.wait_for(app.root,
                           lambda: worker and not worker[0].is_alive(),
                           timeout=60)
                # The release happens on the worker; the dialogs it queued still
                # have to drain before the run is really over.
                h.pump(app.root, 0.5)
                observed["released"] = app.df is None
                observed["exported"] = bool(dialogs["info"])
                observed["errors"] = [t for t, _ in dialogs["error"]]

        # A live mainloop is mandatory: the worker marshals every UI update with
        # root.after(), which raises from a non-Tk thread otherwise.
        h.run_with_mainloop(app.root, scenario)

    h.detail("export completed (success dialog)", observed.get("exported"))
    if observed.get("errors"):
        h.detail("error dialogs", observed["errors"])
    h.detail("app.df is None after the run", observed.get("released"))

    sites = df_clearing_sites()
    for func, lineno in sites:
        h.detail(f"self.df = None at app.py:{lineno}", f"in {func}()")
    releases = [f for f, _ in sites if f != "__init__"]
    h.detail("assignments outside __init__ (i.e. actual releases)",
             releases or "NONE")

    return bool(observed.get("released")), bool(observed.get("exported"))


def main():
    h.require_display()
    h.banner("F-08", "Memory amplification across the export path")

    from flightserver import server_subprocess, connected_connection
    import pyarrow.flight as flight

    s = Sampler()

    with h.isolated_home(), h.temp_dir() as tmp:
        s.take("baseline")

        with server_subprocess(ROWS, COLS) as port:
            h.step(f"Copy 1: real reader.read_all() over Arrow Flight "
                   f"({ROWS:,} x {COLS} strings, server out of process)")
            with h.silence_fd_stderr():
                conn = connected_connection(port)
                options = flight.FlightCallOptions(headers=[conn.bearer_token])
                info = conn.client.get_flight_info(
                    flight.FlightDescriptor.for_command("SELECT *"), options)
                reader = conn.client.do_get(info.endpoints[0].ticket, options)
                table = reader.read_all()

            arrow_bytes = table.nbytes / (1024 * 1024)
            h.detail("rows", f"{table.num_rows:,}")
            h.detail("Arrow buffer size", f"{arrow_bytes:.1f} MB")
            s.take("1. Arrow Table (read_all)")

            h.step("Copies 2+3: real _arrow_to_pandas (cast + to_pandas)")
            df = conn._arrow_to_pandas(table)
            h.detail("DataFrame shape", df.shape)
            s.take("2+3. cast + DataFrame")

            del table
            gc.collect()
            before_export = s.take("after releasing the Arrow table local")

            h.step("The export: real _export_to_excel")
            from openpyxl.workbook.workbook import Workbook
            original_save = Workbook.save
            captured = {}

            def sampling_save(self, filename):
                gc.collect()
                captured["peak"] = h.peak_rss_mb()
                captured["current"] = h.current_rss_mb()
                return original_save(self, filename)

            Workbook.save = sampling_save
            try:
                with h.tk_app(output_dir=tmp, filename="mem.xlsx",
                              autofit=True) as app:
                    app.df = df
                    path = app._export_to_excel()
            finally:
                Workbook.save = original_save

            if "peak" in captured:
                s.add("4+5. rows handed over (pre-save)",
                      captured["peak"], captured["current"])
            s.take("6. XML + zip serialisation")
            file_mb = path.stat().st_size / (1024 * 1024)

        h.step("Measurements")
        s.report()

        base_peak = s.rows[0][1]
        final_peak = s.rows[-1][1]
        peak_growth = final_peak - base_peak
        currents = [c for _, _, c in s.rows if c is not None]
        max_current = max(currents) if currents else None

        h.detail("Arrow source buffers", f"{arrow_bytes:.1f} MB")
        h.detail("final file on disk", f"{file_mb:.1f} MB")
        h.detail("peak RSS growth over baseline", f"{peak_growth:.0f} MB")
        h.detail("=> amplification vs Arrow buffers",
                 f"{peak_growth / arrow_bytes:.0f}x")
        h.detail("=> amplification vs output file", f"{peak_growth / file_mb:.0f}x")

        h.step("The export stage on its own - what this finding is about")
        h.note("Everything before the export is the cost of having the data at "
               "all: the Arrow buffers and the DataFrame. What F-08 is about is "
               "the multiple of that which turning it into a file used to need.")
        export_growth = final_peak - before_export[1]
        export_per_cell = (export_growth * 1024 * 1024) / (ROWS * COLS)
        h.detail("peak before the export", f"{before_export[1]:.0f} MB")
        h.detail("peak after the export", f"{final_peak:.0f} MB")
        h.detail("=> the export stage cost", f"{export_growth:.0f} MB")
        h.detail("=> per cell", f"{export_per_cell:.0f} bytes")
        h.detail("=> budget", f"{EXPORT_BUDGET_BYTES_PER_CELL} bytes per cell")
        within_budget = export_per_cell < EXPORT_BUDGET_BYTES_PER_CELL

        h.step("Concurrent residency vs peak high-water (the audit's caveat)")
        h.detail("max CONCURRENT residency", f"{max_current:.0f} MB")
        h.detail("peak high-water", f"{final_peak:.0f} MB")
        gap = final_peak - max_current
        h.detail("gap", f"{gap:.0f} MB")
        h.note("A small gap means the amplification is genuinely concurrent "
               "residency, not an artifact of summing sequential peaks - i.e. "
               "the copies really are alive at the same time.")

        fidelity_ok = check_fidelity(tmp)
        released, exported = check_frame_released(tmp)

        h.step("Measurement kept for the record: is self_destruct=True defeated?")
        h.note("Unchanged by the fix. Stage 0 established this is single-digit "
               "MB and not where the memory goes; re-measured, not asserted on.")
        plain = self_destruct_probe(
            "no decimal columns (cast is a no-op)",
            pa.table({f"c{i}": pa.array([f"v_{i}_{r}" for r in range(ROWS)])
                      for i in range(COLS)}))
        decimal_held = self_destruct_probe(
            "one decimal128 column (cast materialises)",
            pa.table({
                "amount": pa.array([f"{r}.25" for r in range(ROWS)]).cast(
                    pa.decimal128(20, 2)),
                **{f"c{i}": pa.array([f"v_{i}_{r}" for r in range(ROWS)])
                   for i in range(COLS - 1)},
            }))

        h.step("Extrapolation")
        per_cell = (peak_growth * 1024 * 1024) / (ROWS * COLS)
        projected = (per_cell * 1_000_000 * COLS) / (1024 ** 3)
        h.detail("whole-path cost per cell", f"{per_cell:.0f} bytes")
        h.detail(f"projected peak for 1,000,000 x {COLS}", f"{projected:.1f} GB")

    h.step("Contract check")
    h.detail("1. the export stage stays within budget", within_budget)
    h.detail("2. the exported data is identical to pandas' own", fidelity_ok)
    h.detail("3. the result frame is released when the run ends", released)

    if not exported:
        h.verdict("F-08", h.BLOCKED,
                  "the full Execute run did not reach a success dialog, so the "
                  "release check could not be trusted - fix the run first")
    elif within_budget and fidelity_ok and released:
        h.verdict("F-08", h.NOT_REPRODUCIBLE,
                  f"the export stage costs {export_growth:.0f} MB "
                  f"({export_per_cell:.0f} bytes per cell) to turn a "
                  f"{ROWS * COLS:,}-cell frame into a {file_mb:.1f} MB file, "
                  f"against a budget of {EXPORT_BUDGET_BYTES_PER_CELL} bytes - "
                  f"the workbook is no longer built in memory. Whole-path peak "
                  f"growth is {peak_growth:.0f} MB over {arrow_bytes:.1f} MB of "
                  f"Arrow buffers ({peak_growth / arrow_bytes:.0f}x, was 33x), "
                  f"projecting {projected:.1f} GB for 1M x {COLS} where the "
                  f"audit projected 5.2 GB. Every cell of an 8-dtype frame - "
                  f"nullable Int64, NaN, NaT, bool-with-None, Categorical, "
                  f"Decimal - reads back identical in value and type to what "
                  f"df.to_excel wrote, and app.df is None once the run ends. "
                  f"Unchanged and not where the memory goes: the table local "
                  f"still pins {plain:.1f} MB with no decimal columns and "
                  f"{decimal_held:.1f} MB with a decimal128 column")
    else:
        h.verdict("F-08", h.CONFIRMED,
                  f"export stage {export_growth:.0f} MB = "
                  f"{export_per_cell:.0f} bytes per cell against a "
                  f"{EXPORT_BUDGET_BYTES_PER_CELL}-byte budget "
                  f"(within_budget={within_budget}), data matches pandas="
                  f"{fidelity_ok}, frame released={released}. Whole-path peak "
                  f"growth {peak_growth:.0f} MB over {arrow_bytes:.1f} MB of "
                  f"Arrow buffers = {peak_growth / arrow_bytes:.0f}x, "
                  f"concurrent residency {max_current:.0f} MB against a "
                  f"{final_peak:.0f} MB high-water ({gap:.0f} MB gap), "
                  f"projecting {projected:.1f} GB for 1M x {COLS}")


if __name__ == "__main__":
    main()
