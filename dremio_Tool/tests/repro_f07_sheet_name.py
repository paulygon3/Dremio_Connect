"""
================================================================================
F-07 - Sheet-name rules: not reachable, but the config lies   (Low)
================================================================================
SOURCE-only in AUDIT.md, then executed in Stage 0. This was the one Excel limit
the app was accidentally safe from: the sheet name was the hardcoded literal
'Data', so no user input reached that argument and none of Excel's six
sheet-name rules could be violated.

The hazard was latent rather than live. DEFAULT_CONFIG['output']['sheet_name'] =
'Dremio Data' has been written to config.json since the beginning and never read
back, so a user who edited it saw no effect and no error. That is the "config
lies" half.

The fix makes the setting real, which is the option that resolves the lie rather
than hiding it - the alternative was deleting the key, and a setting that exists
and works is worth more than one that never existed. But AUDIT.md's warning is
the whole point of this script: wiring it up without validation would open all
six rules at once, and openpyxl does NOT enforce them for you.

Measured against openpyxl 25.0.1, it rejects a genuinely empty name and the six
forbidden characters, and ACCEPTS:

    a name longer than 31 characters  (a UserWarning, to a stderr a windowed
                                       PyInstaller build does not have)
    a whitespace-only name            (not empty as far as it is concerned)
    a leading or trailing apostrophe
    'History', which Excel reserves, in any case

Those are the dangerous ones, because the export reports success and Excel then
refuses the file or silently repairs it. So this script checks each rule
individually rather than checking that "a bad name is handled" - and it measures
what openpyxl accepts rather than asserting it, which is how the whitespace-only
case turned up at all.

A bad configured value must not fail the export either. The workbook is fine,
and the user may not know the setting exists - so the contract is: use it when
it is valid, say plainly what was wrong when it is not, and write the default.
================================================================================
"""

REQUIRES_DISPLAY = True

import pandas as pd
from openpyxl import load_workbook

import harness as h

# Every character Excel forbids in a sheet name, plus over-length.
ILLEGAL_NAME = "bad[name]:*?/\\"

# (label, configured value, must be rejected)
RULE_CASES = [
    ("ordinary name", "Dremio Data", False),
    ("padded - trimmed, not rejected", "  Dremio Data  ", False),
    ("exactly 31 characters", "a" * 31, False),
    ("32 characters", "a" * 32, True),
    ("empty", "", True),
    ("whitespace only", "   ", True),
    ("every forbidden character", ILLEGAL_NAME, True),
    ("leading apostrophe", "'quoted", True),
    ("trailing apostrophe", "quoted'", True),
    ("reserved name", "History", True),
    ("reserved name, other case", "hIsToRy", True),
]


def export_with_configured_sheet_name(tmp, configured, filename):
    """Export through the real path and report what was written and said."""
    df = pd.DataFrame({"a": [1]})
    with h.tk_app(output_dir=tmp, filename=filename, autofit=False) as app:
        app.config.set("output", "sheet_name", configured)
        app.config.save_config()
        app.df = df
        path = app._export_to_excel()
        log = app.log_text.get("1.0", "end")
    return load_workbook(path).sheetnames, log


def openpyxl_does_not_enforce_these():
    """
    Re-measure which rules openpyxl actually applies.

    The fix rests on this: if openpyxl enforced everything, the validator would
    be redundant. It does not, and this is what says so rather than a comment.
    """
    h.step("Which rules does openpyxl enforce on its own?")
    from openpyxl import Workbook

    rows = []
    waved_through = []
    for label, value, should_reject in RULE_CASES:
        if not should_reject:
            continue
        wb = Workbook()
        try:
            with h.quiet_stdout():
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    wb.active.title = value
            accepted = True
        except Exception:
            accepted = False
        if accepted:
            waved_through.append(label)
        rows.append([label, repr(value[:20]),
                     "ACCEPTS IT" if accepted else "rejects it"])
    h.table(["invalid name", "value", "openpyxl"], rows)
    h.detail("=> rules openpyxl leaves to the caller", waved_through)
    return waved_through


def validator_covers_every_rule():
    h.step("Does the validator apply every rule?")
    h.add_src_to_path()
    try:
        from utils import validate_sheet_name
    except ImportError:
        # Pre-fix there is no validator, because there is nothing to validate -
        # the setting is never read. That is the finding, not a crash.
        h.detail("sheet-name validator", "ABSENT - no validate_sheet_name")
        return [label for label, _, reject in RULE_CASES if reject]

    rows = []
    wrong = []
    for label, value, should_reject in RULE_CASES:
        ok, error, cleaned = validate_sheet_name(value)
        rejected = not ok
        correct = rejected == should_reject
        if not correct:
            wrong.append(label)
        rows.append([label, repr(value[:22]),
                     "rejected" if rejected else f"accepted -> {cleaned!r}",
                     "ok" if correct else "WRONG"])
    h.table(["rule", "value", "validator says", ""], rows)
    return wrong


def migration_preserves_existing_workbooks(tmp):
    """
    Honouring the setting must not rename the sheet for existing users.

    output.sheet_name has been written to every config.json since the beginning
    and never read, so the value sitting in a real user's file is 'Dremio Data'
    - the old default - while every workbook they have ever produced contains a
    sheet called 'Data'. Simply starting to obey the setting renames it for
    everyone and breaks any formula, Power Query or macro that refers to it.

    So a stored copy of the legacy default is migrated once, on load. It cannot
    be a deliberate choice, because choosing it has never had any effect.

    The migration must be narrow, and all four cases are checked: a fresh
    install, the legacy default, a name the user genuinely chose, and - the one
    that makes the version stamp necessary - the legacy name chosen deliberately
    AFTER the migration has run, which must then stick.
    """
    h.step("Migration: does honouring the setting change anyone's output?")
    h.add_src_to_path()

    import importlib
    import json
    import config as config_module

    # Pre-fix there is no migration and no such constant; the legacy value is
    # simply what DEFAULT_CONFIG still carries. Naming it here lets this section
    # run against that build and report what it does, rather than dying.
    legacy_name = getattr(config_module, "LEGACY_SHEET_NAME", "Dremio Data")

    def manager():
        importlib.reload(config_module)
        return config_module.ConfigManager()

    rows = []

    # The export cases above ran against this same isolated HOME and left their
    # sheet_name in config.json, so "fresh install" has to mean a file that is
    # genuinely not there.
    config_path = manager().config_file
    config_path.unlink(missing_ok=True)

    fresh = manager()
    rows.append(["fresh install", "-", repr(fresh.get("output", "sheet_name")),
                 "Data"])

    config_path.write_text(json.dumps(
        {"output": {"sheet_name": legacy_name}}))
    legacy = manager()
    on_disk = json.loads(config_path.read_text())["output"]["sheet_name"]
    rows.append(["legacy file, old default",
                 repr(legacy_name),
                 repr(legacy.get("output", "sheet_name")), "Data"])

    config_path.write_text(json.dumps({"output": {"sheet_name": "My Sheet"}}))
    custom = manager()
    rows.append(["legacy file, chosen name", repr("My Sheet"),
                 repr(custom.get("output", "sheet_name")), "My Sheet"])

    # Chosen deliberately after migrating. Without the version stamp this would
    # be rewritten on every start, quietly overriding the user forever.
    chosen = manager()
    chosen.set("output", "sheet_name", legacy_name)
    chosen.save_config()
    reopened = manager()
    rows.append(["chosen after migration",
                 repr(legacy_name),
                 repr(reopened.get("output", "sheet_name")),
                 legacy_name])

    for row in rows:
        row.append("ok" if row[2].strip("'\"") == row[3] else "WRONG")
    h.table(["case", "stored", "loaded as", "should be", ""], rows)

    h.detail("migration is written back to disk", on_disk == "Data")
    wrong = [row[0] for row in rows if row[4] == "WRONG"]
    return wrong, on_disk == "Data"


def main():
    h.require_display()
    h.banner("F-07", "Configured sheet_name, and the rules that come with it")

    h.add_src_to_path()
    waved_through = openpyxl_does_not_enforce_these()
    wrong = validator_covers_every_rule()

    with h.isolated_home(), h.temp_dir() as tmp:
        from constants import DEFAULT_CONFIG
        try:
            from constants import DEFAULT_SHEET_NAME
        except ImportError:
            # Pre-fix the name is a literal in app.py rather than a constant.
            DEFAULT_SHEET_NAME = "Data"
        h.detail("DEFAULT_CONFIG['output']['sheet_name']",
                 repr(DEFAULT_CONFIG["output"]["sheet_name"]))

        h.step("A legal configured name: is it actually used?")
        names, log = export_with_configured_sheet_name(
            tmp, "Dremio Data", "legal.xlsx")
        h.detail("configured", repr("Dremio Data"))
        h.detail("sheets written", names)
        legal_used = names == ["Dremio Data"]
        h.detail("the setting reaches the workbook", legal_used)

        h.step("An illegal configured name: refused, reported, export survives")
        names_illegal, log_illegal = export_with_configured_sheet_name(
            tmp, ILLEGAL_NAME, "illegal.xlsx")
        h.detail("configured", repr(ILLEGAL_NAME))
        h.detail("sheets written", names_illegal)
        fell_back = names_illegal == [DEFAULT_SHEET_NAME]
        told = "WARNING" in log_illegal and "sheet name" in log_illegal.lower()
        h.detail("fell back to the default rather than failing", fell_back)
        h.detail("user was told why", told)
        for line in log_illegal.splitlines():
            if "sheet name" in line.lower():
                h.detail("log", line.strip()[:150])

        migration_wrong, migration_persisted = \
            migration_preserves_existing_workbooks(tmp)

    h.step("Contract check")
    h.detail("migration cases wrong", migration_wrong or "none")
    h.detail("migration is persisted", migration_persisted)
    h.detail("rules openpyxl does NOT enforce", waved_through)
    h.detail("validator cases wrong", wrong or "none")
    h.detail("legal configured name is used", legal_used)
    h.detail("illegal name falls back to the default", fell_back)
    h.detail("...and the user is told", told)

    if (not wrong and legal_used and fell_back and told
            and not migration_wrong and migration_persisted):
        h.verdict("F-07", h.NOT_REPRODUCIBLE,
                  f"the setting is real and validated: a legal configured "
                  f"sheet_name now reaches the workbook, so the config no longer "
                  f"lies, and every one of Excel's rules is checked before it "
                  f"gets there - including the {len(waved_through)} openpyxl "
                  f"accepts without complaint ({', '.join(waved_through)}), "
                  f"which are exactly the ones that produce a file Excel refuses "
                  f"or repairs after a successful-looking export. An invalid "
                  f"value does not fail the export: it is reported and "
                  f"{DEFAULT_SHEET_NAME!r} is written instead. Existing users "
                  f"keep the sheet name their workbooks already have - a stored "
                  f"copy of the old default is migrated once on load, since it "
                  f"cannot be a choice when choosing it never did anything, and "
                  f"a version stamp stops that running against a name they pick "
                  f"deliberately afterwards")
    elif not legal_used:
        h.verdict("F-07", h.CONFIRMED,
                  f"the configured sheet_name is still ignored - written to "
                  f"config.json and never read, so the setting is inert "
                  f"(sheets written: {names})")
    else:
        h.verdict("F-07", h.CONFIRMED,
                  f"the setting is read but not handled properly: "
                  f"validator_wrong={wrong} fell_back={fell_back} told={told}")


if __name__ == "__main__":
    main()
