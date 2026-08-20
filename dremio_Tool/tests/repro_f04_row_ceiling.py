"""
================================================================================
F-04 - Row ceiling: export dies with a raw openpyxl error   (Medium)
================================================================================
Tagged EXECUTED + SOURCE. The measured core is the openpyxl boundary and the
header offset; the SOURCE half is the consequence - that the app performs no
row-count check, so the ValueError reaches the generic handler and is shown
verbatim.

The full 1,048,577-row write is deliberately NOT run, exactly as in the audit:
the boundary is provable at the openpyxl cell constructor for a few
microseconds, and the end-to-end write would cost gigabytes to tell us the same
thing.

The guard half used to be checked by grepping the source for a row-ceiling
comparison. That is the same mistake F-22's repro made - a grep reports the
machinery, not the behaviour, and would call a broken guard present. It now
drives the real _export_to_excel with a frame one row over the limit and checks
what actually happens: that it refuses, that the message says something useful,
and that nothing was written.

An over-limit frame is cheap because the guard fires before any write - one
int64 column of 1,048,576 rows is about 8 MB, and to_excel is never reached. The
1,048,576-row *write* is still deliberately NOT run, exactly as in the audit.
================================================================================
"""

REQUIRES_DISPLAY = True

import pandas as pd
from openpyxl import Workbook

import harness as h

SHEET_MAX_ROW = 1048576


def probe_openpyxl_boundary():
    h.step("EXECUTED: openpyxl's own row ceiling, at the cell constructor")
    ws = Workbook().active
    rows = []
    for row in (SHEET_MAX_ROW - 1, SHEET_MAX_ROW, SHEET_MAX_ROW + 1):
        try:
            ws.cell(row=row, column=1)
            rows.append([row, "accepted"])
        except ValueError as e:
            rows.append([row, f"ValueError: {e}"])
    h.table(["row", "result"], rows)
    return rows


def probe_header_offset(tmp):
    h.step("EXECUTED: header offset, via the real _export_to_excel")
    df = pd.DataFrame({"a": [1, 2, 3]})
    with h.tk_app(output_dir=tmp, filename="offset.xlsx", autofit=False) as app:
        app.df = df
        path = app._export_to_excel()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]   # first sheet, whatever it is named
    h.detail("data rows in frame", len(df))
    h.detail("ws.max_row after export", ws.max_row)
    h.detail("header occupies", "sheet row 1; data starts at row 2")
    usable = SHEET_MAX_ROW - (ws.max_row - len(df))
    h.detail("=> true data-row limit", f"{usable:,}")
    return usable


def check_row_guard(tmp, usable):
    """
    Drive the real export one row over the limit and see what happens.

    Three things matter and are checked separately, because a guard can be
    present and still leave the finding intact:

      1. it refuses at all
      2. the message names Excel's limit and what to do - the openpyxl
         ValueError says only "Row numbers must be between 1 and 1048576",
         which mentions neither
      3. nothing was written. openpyxl fails part-way through the write, so the
         unguarded path leaves a partial workbook at the target path (F-11)
    """
    h.step("Does the app refuse an over-limit frame before writing?")

    target = tmp / "over_limit.xlsx"
    over = pd.DataFrame({"a": range(usable + 1)})

    refused = False
    message = ""
    with h.tk_app(output_dir=tmp, filename="over_limit.xlsx",
                  autofit=False) as app:
        app.df = over
        try:
            app._export_to_excel()
            outcome = "exported it"
        except ValueError as e:
            refused = True
            message = str(e)
            outcome = f"refused: {message.splitlines()[0]}"
        except Exception as e:
            outcome = f"{type(e).__name__}: {str(e)[:70]}"

    # A guard that fires but still writes is only half a fix.
    left_behind = target.exists()

    explains = all(term in message for term in ("worksheet", "LIMIT"))
    counts = f"{usable:,}" in message

    h.table(["check", "result"], [
        ["rows offered", f"{usable + 1:,} (one over the limit)"],
        ["outcome", outcome],
        ["message names the worksheet limit and LIMIT", str(explains)],
        ["message states the usable row count", str(counts)],
        ["file left at the target path", str(left_behind)],
    ])
    if message:
        for line in message.splitlines():
            if line.strip():
                h.detail("message", line.strip())

    return refused, explains and counts, left_behind


def check_boundary_is_exact(usable):
    """
    The guard must accept the last legal row and refuse the first illegal one.

    Off-by-one here is the whole finding: the header offset is precisely what
    makes the limit 1,048,575 rather than the 1,048,576 the sheet advertises.
    """
    h.step("Is the guard's boundary the same as openpyxl's?")
    h.add_src_to_path()
    from app import DremioExporter

    # No guard at all is the finding itself, not a reason to crash - this script
    # has to be able to run against a build that has not been fixed.
    guard = getattr(DremioExporter, "_check_row_ceiling", None)
    if guard is None:
        h.detail("row-count guard", "ABSENT - no _check_row_ceiling exists")
        return 0

    import types
    rows = []
    correct = 0
    for count, should_accept in [(usable, True), (usable + 1, False)]:
        stub = types.SimpleNamespace(df=pd.DataFrame({"a": range(count)}))
        try:
            guard(stub)
            accepted = True
        except ValueError:
            accepted = False
        correct += accepted == should_accept
        rows.append([f"{count:,}", "accepted" if accepted else "refused",
                     "ok" if accepted == should_accept else "WRONG"])
    h.table(["data rows", "guard says", ""], rows)
    return correct


def main():
    h.require_display()
    h.banner("F-04", "Excel row ceiling and the raw error the user sees")

    with h.isolated_home(), h.temp_dir() as tmp:
        boundary = probe_openpyxl_boundary()
        usable = probe_header_offset(tmp)
        refused, explains, left_behind = check_row_guard(tmp, usable)
        exact = check_boundary_is_exact(usable)

    rejected = any("ValueError" in str(r[1]) for r in boundary)

    h.step("Contract check")
    h.detail("openpyxl rejects past the ceiling", rejected)
    h.detail("the app refuses before writing", refused)
    h.detail("the message explains the limit and the fix", explains)
    h.detail("a file was left at the target path", left_behind)
    h.detail("guard boundary matches openpyxl", f"{exact} of 2")

    if not rejected:
        h.verdict("F-04", h.NOT_REPRODUCIBLE,
                  "openpyxl did not reject the row past the ceiling, so the "
                  "premise of this finding no longer holds on this version")
    elif refused and explains and not left_behind and exact == 2:
        h.verdict("F-04", h.NOT_REPRODUCIBLE,
                  f"the export refuses an over-limit frame before writing "
                  f"anything: the guard accepts {usable:,} data rows and refuses "
                  f"{usable + 1:,}, matching openpyxl's own boundary once the "
                  f"header offset is taken into account, and no file is left at "
                  f"the target path. The message names the worksheet limit, the "
                  f"usable row count and the LIMIT clause, where openpyxl's own "
                  f"ValueError mentions none of them")
    else:
        h.verdict("F-04", h.CONFIRMED,
                  f"refused={refused} explains={explains} "
                  f"file_left_behind={left_behind} boundary_correct={exact}/2 - "
                  f"true limit is {usable:,} data rows")


if __name__ == "__main__":
    main()
